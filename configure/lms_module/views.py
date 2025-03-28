from django.shortcuts import render
from rest_framework import viewsets, filters
from rest_framework.response import Response
from .models import *
from .serializers import *
from rest_framework import permissions
from user_profile.function_call import *
from user_profile.models import Department
from django.db import IntegrityError            
import random
from django.db.models import *
import ast
from django.template.loader import get_template
from xhtml2pdf import pisa
import time
import pdfkit
from dms_module.serializers import *

class DepartmentAddView(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = GetDepartmentSerializer
    queryset = Department.objects.all().order_by('-id')
    filter_backends = [filters.SearchFilter,filters.OrderingFilter]  # Add the filter backend for search functionality
    ordering_fields = ['department_name']
    search_fields = ['department_name']

    def create(self,request):
            try:
                department_name = request.data.get('department_name')
                department_description = request.data.get('department_description',None)

                if not department_name:
                    return Response({'status': False,'message': 'Department name is required'})

                department_obj = Department.objects.create(department_name=department_name,department_description=department_description)
                department_obj.save()
                return Response({'status': True,'message':"Department created successfully"})
            except Exception as e:
                return Response({"status": False,'message': 'Something went wrong','error': str(e)})
      
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        try:
            if queryset.exists():
                serializer_data = []
                for obj in queryset:
                    context = {'request': request}  # Removed the serial number from the context
                    serializer = GetDepartmentSerializer(obj, context=context)
                    serializer_data.append(serializer.data)

                count = len(serializer_data)
                return Response({
                    "status": True,
                    "message": "Department data fetched successfully",
                    'total_page': 1,  # Since there is no pagination, the total pages will always be 1
                    'total': count,
                    'data': serializer_data
                })
            else:
                return Response({
                    "status": True,
                    "message": "No Department found",
                    "total_page": 0,
                    "total": 0,
                    "data": []
                })
        except Exception as e:
            return Response({"status": False, 'message': 'Something went wrong', 'error': str(e)})


               
class DepartmentUpdatesViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'department_id'

    def update(self, request, *args, **kwargs):
    
        try:
            department_id = self.kwargs.get("department_id")
            department_name = request.data.get('department_name')
            department_description = request.data.get('department_description')
    
            if not Department.objects.filter(id=department_id).exists():
                return Response({"status": False, "message": "Department id not found"})
    
            department_object = Department.objects.get(id=department_id)
            if department_name:
                department_object.department_name = department_name
            if department_description:
                department_object.department_description = department_description
            department_object.save()
    
            return Response({"status": True, "message": "Department updated successfully"})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})    
            
    def destroy(self, request, *args, **kwargs):
        if not request.user.has_perm('dashboard_app.delete_departmentmaster'):
                return Response({"status": False, "message": "You are not authorized to delete department!", "data": []})
        try:
            department_id = request.data.get('department_id')   
            if not Department.objects.filter(id=department_id):
                return Response({"status":False, "message":"Department id not found"})
                     
            department_object = Department.objects.get(id=department_id)
            department_object.delete()
            return Response({"status":True, "message":"Department deleted succesfully"})
        except Exception as e:
                return Response({"status": False,'message': 'Something went wrong','error': str(e)})
            

class PlantAddView(viewsets.ModelViewSet):
    # permission_classes = [permissions.IsAuthenticated]
    serializer_class = GetPlantSerializer
    queryset = Plant.objects.all().order_by('-id')
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    ordering_fields = ['plant_name']
    search_fields = ['plant_name']

    def create(self, request):
        try:
            plant_name = request.data.get('plant_name')
            plant_location = request.data.get('plant_location')
            plant_description = request.data.get('plant_description',"")

            if not plant_name:
                return Response({'status': False, 'message': 'Plant name is required'})
            if not plant_location:
                return Response({'status': False, 'message': 'Plant location is required'})

            plant_obj = Plant.objects.create(plant_name=plant_name, plant_location=plant_location, plant_description=plant_description)
            plant_obj.save()
            return Response({'status': True, 'message': "Plant created successfully"})
        except Exception as e:
            return Response({"status": False, 'message': 'Something went wrong', 'error': str(e)})

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        try:
            if queryset.exists():
                serializer_data = []
                for obj in queryset:
                    context = {'request': request}  # No serial number needed
                    serializer = GetPlantSerializer(obj, context=context)
                    serializer_data.append(serializer.data)

                count = len(serializer_data)
                return Response({
                    "status": True,
                    "message": "Plant data fetched successfully",
                    'total_page': 1, 
                    'total': count,
                    'data': serializer_data
                })
            else:
                return Response({
                    "status": True,
                    "message": "No Plant found",
                    "total_page": 0,
                    "total": 0,
                    "data": []
                })
        except Exception as e:
            return Response({"status": False, 'message': 'Something went wrong', 'error': str(e)})


class PlantUpdatesViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'plant_id'

    def update(self, request, *args, **kwargs):
        # if not request.user.has_perm('dashboard_app.change_plant'):
        #     return Response({"status": False, "message": "You are not authorized to update plant!", "data": []})

        try:
            plant_id = self.kwargs.get("plant_id")
            plant_name = request.data.get('plant_name')
            plant_location = request.data.get('plant_location')
            plant_description = request.data.get('plant_description')

            # Check if the plant exists
            if not Plant.objects.filter(id=plant_id).exists():
                return Response({"status": False, "message": "Plant ID not found"})

            # Fetch the plant object
            plant_object = Plant.objects.get(id=plant_id)

            # Update plant details if provided
            if plant_name:
                plant_object.plant_name = plant_name
            if plant_location:
                plant_object.plant_location = plant_location
            if plant_description:
                plant_object.plant_description = plant_description
            plant_object.save()

            return Response({"status": True, "message": "Plant updated successfully"})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})

    def destroy(self, request, *args, **kwargs):
        if not request.user.has_perm('dashboard_app.delete_plant'):
            return Response({"status": False, "message": "You are not authorized to delete plant!", "data": []})

        try:
            plant_id = request.data.get('plant_id')

            # Check if the plant exists
            if not Plant.objects.filter(id=plant_id).exists():
                return Response({"status": False, "message": "Plant ID not found"})

            # Delete the plant
            plant_object = Plant.objects.get(id=plant_id)
            plant_object.delete()

            return Response({"status": True, "message": "Plant deleted successfully"})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})


class JobRoleAddView(viewsets.ModelViewSet):
    # permission_classes = [permissions.IsAuthenticated]  # Uncomment if authentication is required
    serializer_class = GetJobRoleSerializer
    queryset = JobRole.objects.all().order_by('-id')
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]  # Add search and ordering filters
    ordering_fields = ['job_role_name']  # Fields you want to enable ordering on
    search_fields = ['job_role_name']  # Fields you want to enable search on

    def create(self, request):
        try:
            job_role_name = request.data.get('job_role_name')
            job_role_description = request.data.get('job_role_description')
            plant = request.data.get('plant')
            area = request.data.get('area')
            department = request.data.get('department')

            
            if not job_role_name:
                return Response({'status': False, 'message': 'Job role name is required'})
            if not plant:
                return Response({'status': False, 'message': 'plant is required'})
            if not area:
                return Response({'status': False, 'message': 'area is required'})
            if not department:
                return Response({'status': False, 'message': 'department is required'})
            
            try:
                plant = Plant.objects.get(id=plant)
            except Plant.DoesNotExist:
                return Response({"status": False, "message": "plant not found", "data": []})
 
            try:
                area = Area.objects.get(id=area)
            except Area.DoesNotExist:
                return Response({"status": False, "message": "Area not found", "data": []})
 
            try:
                department = Department.objects.get(id=department)
            except Department.DoesNotExist:
                return Response({"status": False, "message": "Document type not found", "data": []})
 

            job_role_obj = JobRole.objects.create(
                job_role_name=job_role_name,
                job_role_description=job_role_description,
                plant=plant,
                department=department,
                area=area,
            )
            
            job_role_obj.save()

            return Response({'status': True, 'message': "Job role created successfully"})
        except Exception as e:
            return Response({"status": False, 'message': 'Something went wrong', 'error': str(e)})

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        try:
            if queryset.exists():
                serializer_data = []
                for obj in queryset:
                    context = {'request': request}  # No serial number needed
                    serializer = GetJobRoleSerializer(obj, context=context)
                    serializer_data.append(serializer.data)

                count = len(serializer_data)
                return Response({
                    "status": True,
                    "message": "Job role data fetched successfully",
                    'total_page': 1,  # Total pages can always be 1 without pagination
                    'total': count,
                    'data': serializer_data
                })
            else:
                return Response({
                    "status": True,
                    "message": "No Job role found",
                    "total_page": 0,
                    "total": 0,
                    "data": []
                })
        except Exception as e:
            return Response({"status": False, 'message': 'Something went wrong', 'error': str(e)})


class JobRoleUpdatesViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'job_role_id'

    def update(self, request, *args, **kwargs):
        # if not request.user.has_perm('dashboard_app.change_jobrole'):
        #     return Response({"status": False, "message": "You are not authorized to update job role!", "data": []})

        try:
            job_role_id = self.kwargs.get("job_role_id")
            job_role_name = request.data.get('job_role_name')
            job_role_description = request.data.get('job_role_description')

            # Check if the job role exists
            if not JobRole.objects.filter(id=job_role_id).exists():
                return Response({"status": False, "message": "Job role ID not found"})

            # Fetch the job role object
            job_role_object = JobRole.objects.get(id=job_role_id)

            # Update job role details if provided
            if job_role_name:
                job_role_object.job_role_name = job_role_name
            if job_role_description:
                job_role_object.job_role_description = job_role_description
            job_role_object.save()

            return Response({"status": True, "message": "Job role updated successfully"})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})

    def destroy(self, request, *args, **kwargs):
        if not request.user.has_perm('dashboard_app.delete_jobrole'):
            return Response({"status": False, "message": "You are not authorized to delete job role!", "data": []})

        try:
            job_role_id = request.data.get('job_role_id')

            # Check if the job role exists
            if not JobRole.objects.filter(id=job_role_id).exists():
                return Response({"status": False, "message": "Job role ID not found"})

            # Delete the job role
            job_role_object = JobRole.objects.get(id=job_role_id)
            job_role_object.delete()

            return Response({"status": True, "message": "Job role deleted successfully"})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})


class AreaAddView(viewsets.ModelViewSet):
    # permission_classes = [permissions.IsAuthenticated]  # Uncomment if authentication is required
    serializer_class = GetAreaSerializer
    queryset = Area.objects.all().order_by('-id')
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]  # Add search and ordering filters
    ordering_fields = ['area_name']  # Fields you want to enable ordering on
    search_fields = ['area_name', 'department__department_name']  # Fields you want to enable search on, including related department name

    def create(self, request):
        try:
            area_name = request.data.get('area_name')
            department_id = request.data.get('department_id')
            area_description = request.data.get('area_description')

            # Validate required fields
            if not area_name:
                return Response({'status': False, 'message': 'Area name is required'})
            if not department_id:
                return Response({'status': False, 'message': 'Department ID is required'})
            if not area_description:
                return Response({'status': False, 'message': 'Area description is required'})

            # Validate if the department exists
            if not Department.objects.filter(id=department_id).exists():
                return Response({'status': False, 'message': 'Department ID not found'})

            # Create Area object
            area_obj = Area.objects.create(
                area_name=area_name,
                department_id=department_id,
                area_description=area_description
            )
            area_obj.save()

            return Response({'status': True, 'message': "Area created successfully"})
        except Exception as e:
            return Response({"status": False, 'message': 'Something went wrong', 'error': str(e)})

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        try:
            if queryset.exists():
                serializer_data = []
                for obj in queryset:
                    context = {'request': request}  # No serial number needed
                    serializer = GetAreaSerializer(obj, context=context)
                    serializer_data.append(serializer.data)

                count = len(serializer_data)
                return Response({
                    "status": True,
                    "message": "Area data fetched successfully",
                    'total_page': 1,  # Total pages can always be 1 without pagination
                    'total': count,
                    'data': serializer_data
                })
            else:
                return Response({
                    "status": True,
                    "message": "No Area found",
                    "total_page": 0,
                    "total": 0,
                    "data": []
                })
        except Exception as e:
            return Response({"status": False, 'message': 'Something went wrong', 'error': str(e)})


class AreaUpdatesViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'area_id'

    def update(self, request, *args, **kwargs):
        # if not request.user.has_perm('dashboard_app.change_area'):
        #     return Response({"status": False, "message": "You are not authorized to update area!", "data": []})

        try:
            area_id = self.kwargs.get("area_id")
            area_name = request.data.get('area_name')
            department_id = request.data.get('department_id')
            area_description = request.data.get('area_description')

            # Check if the area exists
            if not Area.objects.filter(id=area_id).exists():
                return Response({"status": False, "message": "Area ID not found"})

            # Fetch the area object
            area_object = Area.objects.get(id=area_id)

            # Update area details if provided
            if area_name:
                area_object.area_name = area_name
            if department_id:
                if not Department.objects.filter(id=department_id).exists():
                    return Response({"status": False, "message": "Department ID not found"})
                area_object.department_id = department_id
            if area_description:
                area_object.area_description = area_description
            area_object.save()

            return Response({"status": True, "message": "Area updated successfully"})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})

    def destroy(self, request, *args, **kwargs):
        if not request.user.has_perm('dashboard_app.delete_area'):
            return Response({"status": False, "message": "You are not authorized to delete area!", "data": []})

        try:
            area_id = request.data.get('area_id')

            if not Area.objects.filter(id=area_id).exists():
                return Response({"status": False, "message": "Area ID not found"})

            area_object = Area.objects.get(id=area_id)
            area_object.delete()

            return Response({"status": True, "message": "Area deleted successfully"})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})


class AssessmentViewSet(viewsets.ModelViewSet):
    serializer_class = AssessmentSerializer
    queryset = Assessment.objects.all().order_by('-id')
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    ordering_fields = ['title']
    search_fields = ['title', 'sop_selection', 'assign']  # You can add more fields for searching if needed

    def create(self, request):
        try:
            title = request.data.get('title')
            time_limit = request.data.get('time_limit')
            sop_selection = request.data.get('sop_selection')
            assign = request.data.get('assign')
            pass_percentage = request.data.get('pass_percentage')
            number_of_attempts = request.data.get('number_of_attempts')

            # Validation
            if not title:
                return Response({'status': False, 'message': 'Title is required'})
            if not time_limit:
                return Response({'status': False, 'message': 'Time limit is required'})
            if not sop_selection:
                return Response({'status': False, 'message': 'SOP selection is required'})
            if not assign:
                return Response({'status': False, 'message': 'Assignment (Department/User) is required'})
            if not pass_percentage:
                return Response({'status': False, 'message': 'Pass percentage is required'})
            if not number_of_attempts:
                return Response({'status': False, 'message': 'Number of attempts is required'})

            # Create Assessment
            assessment_obj = Assessment.objects.create(
                title=title,
                time_limit=time_limit,
                sop_selection=sop_selection,
                assign=assign,
                pass_percentage=pass_percentage,
                number_of_attempts=number_of_attempts,
                created_at=timezone.now()
            )
            assessment_obj.save()

            return Response({'status': True, 'message': 'Assessment created successfully'})
        except Exception as e:
            return Response({"status": False, 'message': 'Something went wrong', 'error': str(e)})

    def list(self, request):
        queryset = Assessment.objects.all().order_by('-id')  # Fetching all assessments and ordering them by '-id'
        
        try:
            if queryset.exists():
                serializer = AssessmentSerializer(queryset, many=True)  # Serializing the queryset
                return Response({
                    "status": True,
                    "message": "Assessments fetched successfully",
                    'total': queryset.count(),  # Total number of assessments
                    'data': serializer.data
                })
            else:
                return Response({
                    "status": True,
                    "message": "No Assessments found",
                    "total": 0,
                    "data": []
                })
        except Exception as e:
            return Response({"status": False, 'message': 'Something went wrong', 'error': str(e)})

class AssessmentUpdateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'assessment_id'  # Use this to specify the field used for lookup (e.g., 'assessment_id')

    def update(self, request, *args, **kwargs):
        # Check if the user has the permission to change assessments
        if not request.user.has_perm('app_name.change_assessment'):  # Replace 'app_name' with your app's name
            return Response({"status": False, "message": "You are not authorized to update assessment!", "data": []})

        try:
            # Get the assessment ID from the URL
            assessment_id = self.kwargs.get("assessment_id")
            
            # Get data from the request
            title = request.data.get('title')
            time_limit = request.data.get('time_limit')
            sop_selection = request.data.get('sop_selection')
            assign = request.data.get('assign')
            pass_percentage = request.data.get('pass_percentage')
            number_of_attempts = request.data.get('number_of_attempts')

            # Check if the assessment exists
            if not Assessment.objects.filter(id=assessment_id).exists():
                return Response({"status": False, "message": "Assessment ID not found"})

            # Fetch the assessment object
            assessment_object = Assessment.objects.get(id=assessment_id)

            # Update the assessment details if provided
            if title:
                assessment_object.title = title
            if time_limit:
                assessment_object.time_limit = time_limit
            if sop_selection:
                assessment_object.sop_selection = sop_selection
            if assign:
                assessment_object.assign = assign
            if pass_percentage:
                assessment_object.pass_percentage = pass_percentage
            if number_of_attempts:
                assessment_object.number_of_attempts = number_of_attempts
            
            # Save the updated assessment object
            assessment_object.save()

            return Response({"status": True, "message": "Assessment updated successfully"})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})

class AssessmentQuestionViewSet(viewsets.ModelViewSet):
    serializer_class = AssessmentQuestionSerializer
    queryset = AssessmentQuestion.objects.all().order_by('-id')
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    ordering_fields = ['created_at']
    search_fields = ['sop', 'created_by__full_name', 'department__name']

    def create(self, request):
        try:
            department = request.data.get('department')
            sop = request.data.get('sop')
            questions_data = request.data.get('questions_data')
            marks = request.data.get('marks')
            created_by = request.user  # Assuming user is authenticated

            # Validation
            if not department:
                return Response({'status': False, 'message': 'Department is required'})
            if not questions_data:
                return Response({'status': False, 'message': 'Questions data is required'})

            # Create AssessmentQuestion
            assessment_question = AssessmentQuestion.objects.create(
                department_id=department,
                sop=sop,
                questions_data=questions_data,  # This is a JSON field
                marks=marks,
                created_by=created_by,
                created_at=timezone.now()
            )

            return Response({
                'status': True,
                'message': 'Assessment question created successfully',
                'data': AssessmentQuestionSerializer(assessment_question).data
            })
        except Exception as e:
            return Response({"status": False, 'message': 'Something went wrong', 'error': str(e)})

class AssessmentQuestionUpdateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'assessment_question_id'  # Using a custom lookup field like 'assessment_question_id'

    def update(self, request, *args, **kwargs):
        # Check if the user has the permission to change assessment questions
        if not request.user.has_perm('app_name.change_assessmentquestion'):  # Replace 'app_name' with your app's name
            return Response({"status": False, "message": "You are not authorized to update this assessment question!", "data": []})

        try:
            # Get the assessment question ID from the URL
            assessment_question_id = self.kwargs.get("assessment_question_id")
            
            # Check if the assessment question exists
            if not AssessmentQuestion.objects.filter(id=assessment_question_id).exists():
                return Response({"status": False, "message": "Assessment Question ID not found"})

            # Fetch the assessment question object
            assessment_question = AssessmentQuestion.objects.get(id=assessment_question_id)

            # Get data from the request
            department = request.data.get('department')
            sop = request.data.get('sop')
            questions_data = request.data.get('questions_data')
            marks = request.data.get('marks')

            # Update the assessment question details if provided
            if department:
                assessment_question.department_id = department
            if sop:
                assessment_question.sop = sop
            if questions_data:
                assessment_question.questions_data = questions_data  # Updating the JSON field
            if marks:
                assessment_question.marks = marks
            
            # Save the updated assessment question
            assessment_question.save()

            return Response({
                "status": True,
                "message": "Assessment question updated successfully",
                "data": AssessmentQuestionSerializer(assessment_question).data
            })
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})


class MethodologyCreateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = MethodologySerializer
    queryset = Methodology.objects.all().order_by('-id')
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    ordering_fields = ['methodology_name','created_at']
    search_fields = ['methodology_name', 'created_by__full_name']

    def create(self, request, *args, **kwargs):
        try:
            methodology_name = request.data.get('methodology_name')
            created_by = self.request.user

            # Validation
            if not methodology_name:
                return Response({'status': False, 'message': 'Methodology name is required'})

            # Create Methodology
            methodology = Methodology.objects.create(
                methodology_name=methodology_name,
                created_by=created_by
            )

            serializer = MethodologySerializer(methodology)
            data = serializer.data
            return Response({
                'status': True,
                'message': 'Methodology created successfully',
                'data': data
            })
        except Exception as e:
            return Response({"status": False, 'message': 'Something went wrong', 'error': str(e)})
        
    def list(self, request, *args, **kwargs):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = MethodologySerializer(queryset, many=True)
            data = serializer.data
            return Response({"status": True,"message": "Methodologies fetched successfully","data": data})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        
class MethodologyUpdateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'methodology_id'  # Using a custom lookup field like 'methodology_id'

    def update(self, request, *args, **kwargs):
        # Check if the user has the permission to change methodologies
        # if not request.user.has_perm('lms_module.change_methodology'):
        #     return Response({"status": False, "message": "You are not authorized to update this methodology!", "data": []})
        try:
            # Get the methodology ID from the URL
            methodology_id = self.kwargs.get("methodology_id")
            
            # Check if the methodology exists
            if not Methodology.objects.filter(id=methodology_id).exists():
                return Response({"status": False, "message": "Methodology ID not found"})

            methodology = Methodology.objects.get(id=methodology_id)
            methodology_name = request.data.get('methodology_name')

            if methodology_name:
                methodology.methodology_name = methodology_name

            methodology.save()
            serializer = MethodologySerializer(methodology)
            data = serializer.data
            return Response({
                "status": True,
                "message": "Methodology updated successfully",
                "data": data
            })
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        
    def destroy(self, request, *args, **kwargs):
        # Check if the user has the permission to delete methodologies
        # if not request.user.has_perm('lms_module.delete_methodology'):
        #     return Response({"status": False, "message": "You are not authorized to delete this methodology!", "data": []})
        try:
            # Get the methodology ID from the URL
            methodology_id = self.kwargs.get("methodology_id")
            
            if not Methodology.objects.filter(id=methodology_id).exists():
                return Response({"status": False, "message": "Methodology ID not found"})

            Methodology.objects.filter(id=methodology_id).delete()
            return Response({"status": True, "message": "Methodology deleted successfully"})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        

class TrainingTypeCreateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingTypeSerializer
    queryset = TrainingType.objects.all().order_by('-id')
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    ordering_fields = ['training_type_name','created_at']
    search_fields = ['training_type_name', 'created_by__full_name']

    def create(self, request, *args, **kwargs):
        try:
            training_type_name = request.data.get('training_type_name')
            created_by = self.request.user

            # Validation
            if not training_type_name:
                return Response({'status': False, 'message': 'Training type name is required'})

            # Create Training Type
            training_type = TrainingType.objects.create(
                training_type_name=training_type_name,
                created_by=created_by
            )

            serializer = TrainingTypeSerializer(training_type)
            data = serializer.data
            return Response({
                'status': True,
                'message': 'Training type created successfully',
                'data': data
            })
        except Exception as e:
            return Response({"status": False, 'message': 'Something went wrong', 'error': str(e)})
        
    def list(self, request, *args, **kwargs):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = TrainingTypeSerializer(queryset, many=True)
            data = serializer.data
            return Response({"status": True,"message": "Training types fetched successfully","data": data})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        

class TrainingTypeUpdateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'training_type_id'  # Using a custom lookup field like 'training_type_id'

    def update(self, request, *args, **kwargs):
        try:
            # Get the training type ID from the URL
            training_type_id = self.kwargs.get("training_type_id")
            
            # Check if the training type exists
            if not TrainingType.objects.filter(id=training_type_id).exists():
                return Response({"status": False, "message": "Training type ID not found"})

            training_type = TrainingType.objects.get(id=training_type_id)
            training_type_name = request.data.get('training_type_name')

            if training_type_name:
                training_type.training_type_name = training_type_name

            training_type.save()
            serializer = TrainingTypeSerializer(training_type)
            data = serializer.data
            return Response({
                "status": True,
                "message": "Training type updated successfully",
                "data": data
            })
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        
    def destroy(self, request, *args, **kwargs):
        # Check if the user has the permission to delete training types
        # if not request.user.has_perm('lms_module.delete_trainingtype'):
        #     return Response({"status": False, "message": "You are not authorized to delete this training type!", "data": []})
        try:
            # Get the training type ID from the URL
            training_type_id = self.kwargs.get("training_type_id")
            
            if not TrainingType.objects.filter(id=training_type_id).exists():
                return Response({"status": False, "message": "Training type ID not found"})

            TrainingType.objects.filter(id=training_type_id).delete()
            return Response({"status": True, "message": "Training type deleted successfully"})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        

class TrainingCreateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingCreateSerializer
    queryset = TrainingCreate.objects.all().order_by('-id')
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    ordering_fields = ['training_name','created_at']
    search_fields = ['training_name', 'created_by__full_name']

    def create(self, request, *args, **kwargs):
        try:
            plant = request.data.get('plant')
            training_type = request.data.get('training_type')
            training_number = request.data.get('training_number')
            training_name = request.data.get('training_name')
            training_version = request.data.get('training_version')
            refresher_time = request.data.get('refresher_time')
            training_document = request.data.get('training_document')
            schedule_date = request.data.get('schedule_date')
            methodology_ids_str = request.data.get('methodology_ids',[])
            if not methodology_ids_str:
                return Response({'status': False, 'message': 'Methodology is required'})
            
            try:
                methodology_ids = ast.literal_eval(methodology_ids_str)
                if not isinstance(methodology_ids, list):
                    return Response({'status': False, 'message': 'Methodology must be a valid list'})
                methodology_ids = [int(id) for id in methodology_ids]
            except (ValueError, SyntaxError):
                return Response({'status': False, 'message': 'Methodology must be a valid list of integers'})

            created_by = self.request.user

            # Validation
            if not plant:
                return Response({'status': False, 'message': 'Plant is required'})
            if not training_type:
                return Response({'status': False, 'message': 'Training type is required'})
            if not training_number:
                return Response({'status': False, 'message': 'Training number is required'})
            if not training_name:
                return Response({'status': False, 'message': 'Training name is required'})
            if not training_version:
                return Response({'status': False, 'message': 'Training version is required'})
            if not refresher_time:
                return Response({'status': False, 'message': 'Refresher time is required'})
            if not training_document:
                return Response({'status': False, 'message': 'Training document is required'})
            if not methodology_ids:
                return Response({'status': False, 'message': 'Methodology is required'})

            try:
                plant = Plant.objects.get(id=plant)
            except Plant.DoesNotExist:
                return Response({'status': False, 'message': 'Plant not found'})
            
            try:
                training_type = TrainingType.objects.get(id=training_type)
            except TrainingType.DoesNotExist:
                return Response({'status': False, 'message': 'Traing type not found'})

            # Generate the file path for saving the document
            document_path = get_training_document_upload_path(training_document.name)
            file_path = os.path.join(settings.MEDIA_ROOT, document_path)
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            # Write the document to the generated file path
            with open(file_path, 'wb') as destination:
                for chunk in training_document.chunks():
                    destination.write(chunk)

            # Create Training
            training = TrainingCreate.objects.create(
                plant=plant,
                training_type=training_type,
                training_number=training_number,
                training_name=training_name,
                training_version=training_version,
                refresher_time=refresher_time,
                training_document=training_document,
                created_by=created_by,
                schedule_date=schedule_date
            )
            
            if methodology_ids:
                training.methodology.set(methodology_ids)

            training.save()

            # serializer = TrainingCreateSerializer(training, context = {'request': request})
            # data = serializer.data
            return Response({"status": True,"message": "Training created successfully"})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        

    # def list(self, request, *args, **kwargs):
    #     try:
    #         queryset = self.filter_queryset(self.get_queryset())
    #         serializer = TrainingCreateSerializer(queryset, many=True, context = {'request': request})
    #         data = serializer.data
    #         return Response({"status": True,"message": "Training list fetched successfully","data": data})
    #     except Exception as e:
    #         return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        
    def list(self, request, *args, **kwargs):
        try:
            user = self.request.user
            format_type = DocumentType.objects.filter(document_name="Format").first()
            if user.groups.filter(name="DTC").exists():
                queryset_documents = Document.objects.filter(Q(document_current_status_id=6) | Q(document_current_status_id=7, training_required=True)).exclude(document_type=format_type).exclude(document_current_status=15).exclude(document_current_status=12)

            else:
                job_roles = JobRole.objects.filter(job_assigns__user=user)
                queryset_documents = Document.objects.filter(job_roles__in=job_roles).exclude(document_current_status_id=12).exclude(document_type=format_type).exclude(document_current_status=15).distinct()

            latest_versions = queryset_documents.values('document_title').annotate(
                max_version=Max('version')
            )
            queryset_documents = queryset_documents.filter(
            version=Subquery(
                latest_versions.filter(document_title=OuterRef('document_title')).values('max_version')[:1]
            )
        )
            queryset_documents = self.filter_queryset(queryset_documents)
            document_serializer = DocumentviewSerializer(queryset_documents, many=True, context={'request': request})
            document_data = document_serializer.data

            quiz_sessions = QuizSession.objects.filter(user=user)
            quiz_session_serializer = QuizSessionSerializer(quiz_sessions, many=True, context={'request': request})
            quiz_session_data = quiz_session_serializer.data
    
            # Mapping quiz session data by quiz ID for quick lookup
            quiz_sessions_dict = {}
            for quiz in quiz_session_data:
                quiz_id = quiz["quiz"]
                if quiz_id not in quiz_sessions_dict:
                    quiz_sessions_dict[quiz_id] = []
                quiz_sessions_dict[quiz_id].append(quiz)
    
            # Merging quiz session data into document data
            for document in document_data:
                training_quiz_ids = document.get("training_quiz_ids", [])
                document["quiz_sessions"] = []
    
                for quiz_id in training_quiz_ids:
                    if quiz_id in quiz_sessions_dict:
                        document["quiz_sessions"].extend(quiz_sessions_dict[quiz_id])
    
            response_data = {
                "documents": document_data,
            }
    
            return Response({"status": True, "message": "Document list fetched successfully", "document_data": response_data})
        
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})



class TrainingUpdateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingCreateSerializer
    queryset = TrainingCreate.objects.all().order_by('-id')
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    ordering_fields = ['training_name', 'created_at']
    search_fields = ['training_name', 'created_by__full_name']
    lookup_field = 'training_id'

    # Override the default `update` method provided by ModelViewSet
    def update(self, request, *args, **kwargs):
        try:
            training_id = self.kwargs.get("training_id")

            training = TrainingCreate.objects.get(id=training_id)
            if not training:
                return Response({"status": False, "message": "Training ID not found", "data": []})
            
            plant = request.data.get('plant', training.plant.id)
            training_type = request.data.get('training_type', training.training_type.id)
            training_number = request.data.get('training_number', training.training_number)
            training_name = request.data.get('training_name', training.training_name)
            training_version = request.data.get('training_version', training.training_version)
            refresher_time = request.data.get('refresher_time', training.refresher_time)
            training_document = request.FILES.get('training_document', training.training_document)  # Use request.FILES for uploaded files
            methodology_ids_str = request.data.get('methodology_ids', None)
            training_status = request.data.get('training_status', training.training_status)
            schedule_date = request.data.get('schedule_date', training.schedule_date)
            
            try:
                plant = Plant.objects.get(id=plant)
            except Plant.DoesNotExist:
                return Response({'status': False, 'message': 'Plant not found'})
            
            try:
                training_type = TrainingType.objects.get(id=training_type)
            except TrainingType.DoesNotExist:
                return Response({'status': False, 'message': 'Traing type not found'})
            
            try:
                methodology_ids = ast.literal_eval(methodology_ids_str)
                if not isinstance(methodology_ids, list):
                    return Response({'status': False, 'message': 'Methodology must be a valid list'})
                methodology_ids = [int(id) for id in methodology_ids]
            except (ValueError, SyntaxError):
                return Response({'status': False, 'message': 'Methodology must be a valid list of integers'})
            
            if training_name:
                training.training_name = training_name
            if training_version:
                training.training_version = training_version
            if refresher_time:
                training.refresher_time = refresher_time
            if plant:
                training.plant = plant
            if training_type:
                training.training_type = training_type
            if training_number:
                training.training_number = training_number
            if training_status:
                training.training_status = training_status
            if schedule_date:
                training.schedule_date = schedule_date

            training.training_updated_at = timezone.now()

            if training_document:
                document_path = get_training_document_upload_path(training_document.name)
                file_path = os.path.join(settings.MEDIA_ROOT, document_path)
                os.makedirs(os.path.dirname(file_path), exist_ok=True)

                with open(file_path, 'wb') as destination:
                    for chunk in training_document.chunks():
                        destination.write(chunk)

                training.training_document = document_path

            if methodology_ids:
                training.methodology.set(methodology_ids)

            training.save()
            serializer = TrainingCreateSerializer(training, context={'request': request})
            data = serializer.data
            return Response({
                "status": True,
                "message": "Training updated successfully",
                "data": data
            })

        except TrainingCreate.DoesNotExist:
            return Response({"status": False, "message": "Training not found", "data": []})
        except Exception as e:
            return Response({"status": False, "message": f"Something went wrong: {str(e)}", "data": []})

class TrainingStatusUpdateViewset(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingStatusSerializer
    queryset = TrainingCreate.objects.all().order_by('-id')
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    ordering_fields = ['training_name', 'created_at']

    def update(self, request, *args, **kwargs):
        try:
            training_id = self.kwargs.get("training_id")
            training = TrainingCreate.objects.get(id=training_id)
            if not training:
                return Response({"status": False, "message": "Training ID not found", "data": []})
            
            training_status = request.data.get('training_status')
            if not training_status:
                return Response({"status": False, "message": "Training status is required", "data": []})
            
            training.training_status = training_status
            training.save()
            serializer = TrainingStatusSerializer(training, context={'request': request})
            data = serializer.data
            return Response({
                "status": True,
                "message": "Training status updated successfully",
                "data": data
            })

        except TrainingCreate.DoesNotExist:
            return Response({"status": False, "message": "Training not found", "data": []})
        
class TrainingSectionViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingSectionSerializer
    queryset = TrainingSection.objects.all().order_by('-id')
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['training']

    def create(self, request, *args, **kwargs):
        try:
            user = self.request.user
            # training_id = request.data.get('training_id')
            # training = TrainingCreate.objects.get(id=training_id)
            document_id = request.data.get('document_id')
            document = Document.objects.get(id=document_id)
            # if not training:
            #     return Response({"status": False, "message": "Training ID not found", "data": []})
            if not document:
                return Response({"status": False, "message": "Document ID not found", "data": []})
            
            section_name = request.data.get('section_name')
            section_description = request.data.get('section_description')
            section_order = request.data.get('section_order')
            if not section_name:
                return Response({"status": False, "message": "Section name is required", "data": []})
            if section_order is not None:
                section_order = str(section_order)

            training_section = TrainingSection.objects.create(
                # training=training,
                document=document,
                
                section_name=section_name,
                section_description=section_description,
                section_order=section_order,
                created_by=user
            )

            serializer = TrainingSectionSerializer(training_section, context = {'request': request})
            data = serializer.data
            return Response({"status": True,"message": "Training section created successfully","data": data})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        

    def list(self, request, *args, **kwargs):
        try:
            # training_id = request.query_params.get('training_id')
            document_id = request.query_params.get('document_id')
            # if not training_id:
            #     return Response({"status": False, "message": "training_id is required", "data": []})
            if not document_id:
                return Response({"status": False, "message": "document_id is required", "data": []})
            
            
            # queryset = TrainingSection.objects.filter(training_id=training_id)
            queryset = TrainingSection.objects.filter(document_id=document_id)

            if queryset.exists():
                serializer = TrainingSectionSerializer(queryset, many=True, context={'request': request})
                return Response({"status": True, "message": "Training section list", "data": serializer.data})
            else:
                return Response({"status": True, "message": "No training section found", "data": []})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        

class TrainingSectionUpdateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingSectionSerializer
    queryset = TrainingSection.objects.all().order_by('-id')
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['training']
    lookup_field = 'training_section_id' 
    def update(self, request, *args, **kwargs):
        try:
            user = self.request.user
            training_section_id = self.kwargs.get("training_section_id")
            if not training_section_id:
                return Response({"status": False, "message": "Section ID is required", "data": []})
            
            try:
                section = TrainingSection.objects.get(id=training_section_id)
            except TrainingSection.DoesNotExist:
                return Response({"status": False, "message": "Section ID not found", "data": []})
            
            section_name = request.data.get('section_name', section.section_name)
            section_description = request.data.get('section_description', section.section_description)
            section_order = request.data.get('section_order', section.section_order)
            reason_for_update = request.data.get('reason_for_update', section.reason_for_update)
            training_section_active_status = request.data.get('training_section_active_status', section.training_section_active_status)

            if section_order is not None:
                section_order = str(section_order)

            if section_name:
                section.section_name = section_name
            if section_description:
                section.section_description = section_description
            if section_order:
                section.section_order = section_order
            if reason_for_update:
                section.reason_for_update = reason_for_update

            if training_section_active_status is not None:
                if isinstance(training_section_active_status, bool):
                    section.training_section_active_status = training_section_active_status
                else:
                    return Response({"status": False, "message": "Invalid value for 'training_section_active_status'. It must be a boolean value (True/False).", "data": []})

            section.updated_by = user
            section.save()

            serializer = TrainingSectionSerializer(section, context={'request': request})
            data = serializer.data
            return Response({"status": True, "message": "Training section updated successfully", "data": data})

        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "data": []})
        


    def destroy(self, request, *args, **kwargs):
        try:
            training_section_id = self.kwargs.get("training_section_id")
            if not training_section_id:
                return Response({"status": False, "message": "Section ID is required", "data": []})
            
            try:
                training_section_id = TrainingSection.objects.get(id=training_section_id)
            except TrainingSection.DoesNotExist:
                return Response({"status": False, "message": "Section ID not found", "data": []})
            
            materials = training_section_id.materials.all()
            for material in materials:
                material.material_file.clear() 
                material.delete()

            training_section_id.delete()

            return Response({"status": True, "message": "Training section and associated materials deleted successfully", "data": []})

        except Exception as e:
            return Response({"status": False, "message": f"Something went wrong: {str(e)}", "data": []})
                
from django.db.models import Prefetch

class TrainingMaterialCreateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingNestedSectionSerializer
    queryset = TrainingMaterial.objects.all().order_by('-material_created_at')

    def create(self, request, *args, **kwargs):
        try:
            user = self.request.user
            section_id = request.data.get('section_ids')

            if not section_id:
                return Response({"status": False, "message": "Section IDs are required", "data": []})

            sections = TrainingSection.objects.get(id=section_id)
            if not sections:
                return Response({"status": False, "message": "Section IDs not found", "data": []})

            material_title = request.data.get('material_title')
            material_type = request.data.get('material_type')
            material_files = request.FILES.getlist('material_file')  # Get the list of files
            minimum_reading_time = request.data.get('minimum_reading_time')

            if not material_title:
                return Response({"status": False, "message": "Material title is required", "data": []})
            if not material_type:
                return Response({"status": False, "message": "Material type is required", "data": []})
            if material_type not in dict(TrainingMaterial.MATERIAL_CHOICES).keys():
                return Response({"status": False, "message": "Invalid material type", "data": []})
            if not material_files:
                return Response({"status": False, "message": "Material file is required", "data": []})

            # Create the TrainingMaterial object
            training_material = TrainingMaterial.objects.create(
                material_title=material_title,
                section=sections,
                material_type=material_type,
                minimum_reading_time=minimum_reading_time,
                created_by=user,
                material_created_at=timezone.now()
            )
            for material_file in material_files:
                attachment = TrainingMaterialAttachments.objects.create(
                    user=user,
                    material_file=material_file
                )
                training_material.material_file.add(attachment)
            training_material.save()

            return Response({"status": True, "message": "Training material created successfully", "data": []})

        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e), "data": []})
        

    def list(self, request, *args, **kwargs):
        try:
            training_materials = Prefetch('materials', queryset=TrainingMaterial.objects.all().order_by('-id'))
            sections = TrainingSection.objects.prefetch_related(training_materials).all().order_by('-id')
            serializer = TrainingNestedSectionSerializer(sections, many=True, context={'request': request})
            data = serializer.data
            return Response({"status": True,"message": "Training section list fetched successfully","data": data})
        except Exception as e:
            return Response({"status": False,"message": "Something went wrong","error": str(e)})
        
class TrainingIdWiseTrainingSectionViewset(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingSectionSerializer

    def list(self, request, *args, **kwargs):
        try:
            # training_id = request.query_params.get('training_id')
            # if not training_id:
                # return Response({"status": False, "message": "Training ID is required", "data": []})
            document_id = request.query_params.get('document_id')
            if not document_id:
                return Response({"status": False, "message": "Document ID is required", "data": []})

            # queryset = TrainingSection.objects.filter(training=training_id)
            queryset = TrainingSection.objects.filter(document=document_id)
            
            serializer = TrainingSectionSerializer(queryset, many=True, context = {'request': request})
            data = serializer.data
            return Response({"status": True,"message": "Training section list fetched successfully","data": data})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        

class TrainingSectionWiseTrainingMaterialViewset(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingMaterialSerializer

    def list(self, request, *args, **kwargs):
        try:
            section_id = request.query_params.get('section_id')
            if not section_id:
                return Response({"status": False, "message": "Section ID is required", "data": []})
            
            section = TrainingSection.objects.filter(id=section_id).first()
            if not section:
                return Response({"status": False, "message": "Section not found", "data": []})

            queryset = TrainingMaterial.objects.filter(section=section_id)
            serializer = TrainingMaterialSerializer(queryset, many=True, context = {'request': request})
            section_data = {
                "id": section.id,
                "section_name": section.section_name,
                "section_description": section.section_description,
                "section_order": section.section_order,
                "training_name": section.document.document_title,
            }
            return Response({"status": True,"message": "Training material list fetched successfully","data": {"section": section_data,"materials": serializer.data}})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        

class TrainingMaterialUpdateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingMaterialSerializer
    queryset = TrainingMaterial.objects.all().order_by('-material_created_at')
    lookup_field = 'training_material_id'

    def update(self, request, *args, **kwargs):
        try:
            user = self.request.user
            training_material_id = self.kwargs.get('training_material_id')

            if not training_material_id:
                return Response({"status": False, "message": "Training material ID is required", "data": []})

            try:
                training_material = TrainingMaterial.objects.get(id=training_material_id)
            except TrainingMaterial.DoesNotExist:
                return Response({"status": False, "message": "Training material not found", "data": []})

            section_id = request.data.get('section_id', None)
            if section_id:
                try:
                    section = TrainingSection.objects.get(id=section_id)
                    training_material.section = section
                except TrainingSection.DoesNotExist:
                    return Response({"status": False, "message": "Section not found", "data": []})

            material_title = request.data.get('material_title', training_material.material_title)
            material_type = request.data.get('material_type', training_material.material_type)
            material_files = request.FILES.getlist('material_file', [])
            minimum_reading_time = request.data.get('minimum_reading_time', training_material.minimum_reading_time)

            if material_title:
                training_material.material_title = material_title
            if material_type:
                training_material.material_type = material_type
            if minimum_reading_time:
                training_material.minimum_reading_time = minimum_reading_time

            training_material.updated_by = user
            training_material.material_updated_at = timezone.now()

            training_material.save()
            if material_files:
                training_material.material_file.clear()

                for material_file in material_files:
                    attachment = TrainingMaterialAttachments.objects.create(
                        user=user,
                        material_file=material_file
                    )
                    training_material.material_file.add(attachment)

            training_material.save()
            serializer = TrainingMaterialSerializer(training_material, context={'request': request})
            data = serializer.data

            return Response({"status": True, "message": "Training material updated successfully", "data": data})

        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e), "data": []})

    
    def destroy(self, request, *args, **kwargs):
        try:
            training_material_id = self.kwargs.get('training_material_id')
            try:
                training_material_id = TrainingMaterial.objects.get(id=training_material_id)
            except TrainingMaterial.DoesNotExist:
                return Response({"status": False, "message": "Section ID not found", "data": []})
            
            training_material_id.delete()
            return Response({"status": True, "message": "Material deleted successfully", "data": []})

        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "data": []})
    
    
class TrainingQuestionCreateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingQuestinSerializer
    queryset = TrainingQuestions.objects.all().order_by('-question_created_at')
    
    def create(self, request, *args, **kwargs):
        try:
            user = self.request.user
            # training = request.data.get('training_id')
            # if not training:
                # return Response({"status": False, "message": "Training ID is required", "data": []})
            document_id = self.request.POST.get('document_id')
            if not document_id:
                return Response({"status": False, "message": "Document ID is required", "data": []})

            # training = TrainingCreate.objects.get(id=training)
            # if not training:
                # return Response({"status": False, "message": "Training ID not found", "data": []})
            document = Document.objects.get(id=document_id)
            if not document:
                return Response({"status": False, "message": "Document ID not found", "data": []})

            question_type = request.data.get('question_type')
            question_text = request.data.get('question_text')
            options = request.data.get('options', [])
            correct_answer = request.data.get('correct_answer')
            marks = request.data.get('marks')

            if not question_type:
                return Response({"status": False, "message": "Question type is required", "data": []})
            if not question_text:
                return Response({"status": False, "message": "Question text is required", "data": []})
            
            if question_type == 'mcq' and not options:
                return Response({"status": False, "message": "MCQ questions must have options", "data": []})

            if question_type == 'mcq' and correct_answer not in options:
                return Response({"status": False, "message": "MCQ correct answer must be one of the options", "data": []})

            if question_type == 'true_false' and correct_answer not in ['True', 'False']:
                return Response({"status": False, "message": "True/False correct answer must be 'True' or 'False'", "data": []})

            if question_type == 'fill_in_the_blank' and not correct_answer:
                return Response({"status": False, "message": "Fill-in-the-blank questions must have a correct answer", "data": []})
            selected_file_type = request.data.get('selected_file_type')
            selected_file = request.FILES.get('selected_file', None)

            training_question = TrainingQuestions.objects.create(
                # training=training,
                document=document,
                
                question_type=question_type,
                question_text=question_text,
                options=options,
                correct_answer=correct_answer,
                marks=marks,
                created_by=user,
                selected_file_type = selected_file_type,
                selected_file=selected_file,
                question_created_at=timezone.now()
            )

            # Serialize and return the response
            serializer = TrainingQuestinSerializer(training_question, context={'request': request})
            data = serializer.data
            return Response({"status": True, "message": "Training question created successfully", "data": data})

        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e), "data": []})
        

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = TrainingQuestinSerializer(queryset, many=True, context = {'request': request})
            data = serializer.data
            return Response({"status": True,"message": "Training question list fetched successfully","data": data})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})


class TrainingIdWiseQuestionsViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingQuestinSerializer
    queryset = TrainingQuestions.objects.all().order_by('-question_created_at')
    # lookup_field = 'training_id'
    lookup_field = 'document_id'


    def list(self, request, *args, **kwargs):
        try:
            # training_id = self.kwargs.get('training_id')
            document_id = self.kwargs.get('document_id')
            # if not training_id:
            #     return Response({
            #         "status": False,
            #         "message": "Training ID is required.",
            #         "data": []
            #     })
            if not document_id:
                return Response({"status": False, "message": "Document ID is required", "data": []})
                    

            # queryset = self.filter_queryset(self.get_queryset().filter(training_id=training_id))
            queryset = self.filter_queryset(self.get_queryset().filter(document_id=document_id))
            serializer = TrainingQuestinSerializer(queryset, many=True, context={'request': request})
            data = serializer.data

            return Response({
                "status": True,
                "message": "Training question list fetched successfully",
                "data": data
            })

        except Exception as e:
            return Response({
                "status": False,
                "message": "Something went wrong",
                "error": str(e)
            })


class TrainingQuestionUpdateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingQuestinSerializer
    queryset = TrainingQuestions.objects.all().order_by('-question_created_at')
    lookup_field = 'training_question_id'

    
    def update(self, request, *args, **kwargs):
        try:
            user = self.request.user
            training_question_id = self.kwargs.get('training_question_id')  # Fix to match the URL
            if not training_question_id:
                return Response({"status": False, "message": "Training question ID is required", "data": []})

            # Fetch the TrainingQuestions object
            section = TrainingQuestions.objects.get(id=training_question_id)
            if not section:
                return Response({"status": False, "message": "Training question ID not found", "data": []})

            # Get data from the request or use existing values (for updates)
            question_type = request.data.get('question_type', section.question_type)
            question_text = request.data.get('question_text', section.question_text)
            options = request.data.get('options', section.options)
            correct_answer = request.data.get('correct_answer', section.correct_answer)
            marks = request.data.get('marks', section.marks)
            status = request.data.get('status', section.status)
            # image_file = request.FILES.get('image_file', section.image_file)
            # audio_file = request.FILES.get('audio_file', section.audio_file)
            # video_file = request.FILES.get('video_file', section.video_file)
            if isinstance(status, str):
                status = status.lower() in ['true', '1']
            # Validation checks for question_type and other fields
            if not question_type:
                return Response({"status": False, "message": "Question type is required", "data": []})
            if not question_text:
                return Response({"status": False, "message": "Question text is required", "data": []})

            # Additional validation for specific question types
            if question_type == 'mcq' and not options:
                return Response({"status": False, "message": "MCQ questions must have options", "data": []})

            if question_type == 'mcq' and correct_answer not in options:
                return Response({"status": False, "message": "MCQ correct answer must be one of the options", "data": []})

            if question_type == 'true_false' and correct_answer not in ['True', 'False']:
                return Response({"status": False, "message": "True/False correct answer must be 'True' or 'False'", "data": []})

            if question_type == 'fill_in_the_blank' and not correct_answer:
                return Response({"status": False, "message": "Fill-in-the-blank questions must have a correct answer", "data": []})

            # Update the fields of the section
            section.question_type = question_type
            section.question_text = question_text
            section.options = options
            section.correct_answer = correct_answer
            section.marks = marks
            section.status = status

            # Only update the audio or video file if they are provided in the request
            # if image_file:
            #     section.image_file = image_file
                
            # if audio_file:
            #     section.audio_file = audio_file

            # if video_file:
            #     section.video_file = video_file

            section.updated_by = user
            section.question_updated_at = timezone.now()
            section.save()

            # Serialize and return the updated data
            serializer = TrainingQuestinSerializer(section, context={'request': request})
            data = serializer.data
            return Response({"status": True, "message": "Training question updated successfully", "data": data})

        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e), "data": []})
        

    def destroy(self, request, *args, **kwargs):
        try:
            training_question_id = self.kwargs.get('training_question_id')
            try:
                section = TrainingQuestions.objects.get(id=training_question_id)
            except TrainingQuestions.DoesNotExist:
                return Response({"status": False, "message": "Question ID not found", "data": []})
            
            section.delete()
            return Response({"status": True, "message": "Training question deleted successfully", "data": []})

        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "data": []})


class ActiveDeactiveQuestionViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingQuestinSerializer
    lookup_field = 'question_id'

    def update(self, request, *args, **kwargs):
        try:
            question_id = self.kwargs.get('question_id')
            if not question_id:
                return Response({"status": False, "message": "Question ID is required", "data": []})
            
            question = TrainingQuestions.objects.get(id=question_id)
            if not question:
                return Response({"status": False, "message": "Question ID not found", "data": []})
            
            if question.status == True:
                question.status = False
                question.save()
                return Response({"status": True, "message": "Question deactivated successfully", "data": []})
            question.status = True
            question.save()
            return Response({"status": True, "message": "Question activated successfully", "data": []})
        
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e), "data": []})


import json
import ipdb
class TrainingQuizCreateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingQuizSerializer
    queryset = TrainingQuiz.objects.all().order_by('-created_at')

    def create(self, request, *args, **kwargs):
        try:
            # Extracting data from the request
            user = self.request.user
            # training_id = request.data.get('training_id')
            document_id = request.data.get('document_id')
            quiz_name = request.data.get('name')
            pass_criteria = request.data.get('pass_criteria')
            quiz_time = request.data.get('quiz_time')
            quiz_type = request.data.get('quiz_type')
            total_marks = int(request.data.get('total_marks', 0)) 
            marks_breakdown = request.data.get('marks_breakdown')  # e.g., {'1': 5, '2': 3, '3': 2}
            selected_questions = request.data.get('selected_questions', [])  # Only relevant for manual quizzes
            pass_criteria = int(pass_criteria) if pass_criteria else 0
            if pass_criteria > total_marks:
                return Response({"status": False, "message": "Pass criteria cannot be greater than total marks", "data": []})
            # Validate required fields
            if not all([document_id, quiz_name, pass_criteria, quiz_time, quiz_type, total_marks]):
                return Response({"status": False, "message": "Missing required fields", "data": []})

            try:
                # training = TrainingCreate.objects.get(id=training_id)
                document = Document.objects.get(id=document_id)
            # except TrainingCreate.DoesNotExist:
            #     return Response({"status": False, "message": "Training not found", "data": []})
            except Document.DoesNotExist:
                return Response({"status": False, "message": "Document not found", "data": []})

            # Parse marks_breakdown if it's a string (in case it's sent as a string representation)
            if isinstance(marks_breakdown, str):
                try:
                    marks_breakdown = json.loads(marks_breakdown)  # Convert string to dictionary
                except json.JSONDecodeError:
                    return Response({"status": False, "message": "Invalid marks_breakdown format", "data": []})

            # Create the new quiz
            quiz = TrainingQuiz.objects.create(
                quiz_name=quiz_name,
                pass_criteria=pass_criteria,
                quiz_time=quiz_time,
                quiz_type=quiz_type,
                created_by=user,
                # training=training,
                document=document,
            )

            total_marks_accumulated = 0  
            total_questions = 0  

            # # Handle auto-type quizzes
            if quiz_type == 'auto':
                # Marks breakdown is a dictionary, iterate through it
                for marks, count in marks_breakdown.items():
                    marks = int(marks)  # Ensure marks is an integer
                    count = int(count)  # Ensure count is an integer

                    questions = list(TrainingQuestions.objects.filter(
                        # training=training,  # The training filter
                        document=document,
                        marks=marks,        # Marks filter
                        status=True          # Only active questions
                    ))

                    if len(questions) < count:
                        return Response({"status": False,"message": f"Not enough questions with {marks} marks. Found {len(questions)} questions.","data": []})

                    random.shuffle(questions)
                    # Select the required number of questions
                    # selected_questions = random.sample(questions, count)
                    selected_questions = questions  # Select all available questions
                    

                    potential_marks = total_marks_accumulated + (marks * count)
                    if potential_marks > total_marks:
                        return Response({"status": False,"message": f"Total marks exceeded. The selected questions' marks total {potential_marks}, which exceeds the input total_marks of {total_marks}.","data": []})

                    # Create QuizQuestion for each selected question
                    for question in selected_questions:
                        QuizQuestion.objects.create(quiz=quiz, question=question, marks=marks)

                    total_marks_accumulated += marks * count
                    total_questions += count

            

            # Handle manual-type quizzes
            elif quiz_type == 'manual':
                # Validate selected_questions for manual quiz creation
                if not selected_questions or not isinstance(selected_questions, list):
                    return Response({
                        "status": False,
                        "message": "You must provide a list of selected questions for manual quiz creation.",
                        "data": []
                    })

                # Ensure selected_questions is always a list of integers (question IDs)
                if isinstance(selected_questions, str) and selected_questions.strip() == "":
                    selected_questions = []  # Handle the case where it's an empty string

                if not isinstance(selected_questions, list):
                    return Response({
                        "status": False,
                        "message": "selected_questions must be a list.",
                        "data": []
                    })

                # Validate that each item in selected_questions is an integer (question ID)
                if any(not isinstance(q, int) for q in selected_questions):
                    return Response({
                        "status": False,
                        "message": "Each element in selected_questions must be an integer (question ID).",
                        "data": []
                    })

                questions = TrainingQuestions.objects.filter(
                    id__in=selected_questions,  
                    # training=training,  
                    document=document,  
                    status=True
                )

                if len(questions) != len(selected_questions):
                    return Response({
                        "status": False,
                        "message": "Some of the selected questions are invalid or inactive.",
                        "data": []
                    })

                random.shuffle(questions)
                # Create QuizQuestion for each selected question
                for question in questions:
                    if total_marks_accumulated + question.marks > total_marks:
                        return Response({
                            "status": False,
                            "message": f"Adding this question would exceed the total marks. Current total: {total_marks_accumulated}, question marks: {question.marks}",
                            "data": []
                        })
                    QuizQuestion.objects.create(quiz=quiz, question=question, marks=question.marks)
                    total_marks_accumulated += question.marks
                    total_questions += 1

            # Check for total marks mismatch
            if total_marks_accumulated != total_marks:
                return Response({
                    "status": False,
                    "message": f"Total marks mismatch. The accumulated marks are {total_marks_accumulated}, but the input total_marks was {total_marks}. Please adjust.",
                    "data": []
                })

            # Save the quiz with final total marks and total questions
            quiz.total_marks = total_marks_accumulated
            quiz.total_questions = total_questions
            quiz.save()

            # Return the quiz data
            serializer = TrainingQuizSerializer(quiz, context={'request': request})
            return Response({"status": True, "message": "Quiz created successfully", "data": serializer.data})

        except IntegrityError as e:
            return Response({"status": False, "message": "Database Integrity Error", "error": str(e), "data": []})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e), "data": []})







class TrainingQuizList(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'document_id'
    def list(self, request, *args, **kwargs):
        try:
            document_id = kwargs.get('document_id')
            document = Document.objects.get(id=document_id)
            queryset = list(TrainingQuiz.objects.filter(document=document))
            random.shuffle(queryset)
            serializer = TrainingQuizSerializer(queryset, many=True, context={'request': request})
            return Response({"status": True, "message": "Quizzes retrieved successfully", "data": serializer.data})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e), "data": []})



class TrainingQuizUpdateView(viewsets.ModelViewSet):
    queryset = TrainingQuiz.objects.all()
    serializer_class = TrainingQuizSerializer
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'training_quiz_id'

    def update(self, request, *args, **kwargs):
        try:
            user = self.request.user
            training_quiz_id = self.kwargs.get('training_quiz_id')
            quiz = TrainingQuiz.objects.get(id=training_quiz_id)
            # Check if the quiz status is False user can not update
            if not quiz.status:
                return Response({"status": False, "message": "Quiz is not active", "data": []})

            # training_id = request.data.get('training_id', quiz.training.id)
            # document_id = request.data.get('document_id', quiz.document.id)
            quiz_name = request.data.get('name', quiz.quiz_name)
            pass_criteria = request.data.get('pass_criteria', quiz.pass_criteria)
            quiz_time = request.data.get('quiz_time', quiz.quiz_time)
            quiz_type = request.data.get('quiz_type', quiz.quiz_type)
            total_marks = int(request.data.get('total_marks', quiz.total_marks)) 
            marks_breakdown = request.data.get('marks_breakdown')  # e.g., {'1': 5, '2': 3, '3': 2}
            selected_questions = request.data.get('selected_questions', [])
            status = request.data.get('status',quiz.status)

            if not all([quiz_name, pass_criteria, quiz_time, quiz_type, total_marks]):
                return Response({"status": False, "message": "Missing required fields", "data": []})

            # try:
                # training = TrainingCreate.objects.get(id=training_id)
                # document = Document.objects.get(id=document_id)
            # except TrainingCreate.DoesNotExist:
            #     return Response({"status": False, "message": "Training not found", "data": []})
            # except Document.DoesNotExist:
            #     return Response({"status": False, "message": "Document not found", "data": []})

            # quiz.training = training
            # quiz.document = document
            quiz.quiz_name = quiz_name
            quiz.pass_criteria = pass_criteria
            quiz.quiz_time = quiz_time
            quiz.quiz_type = quiz_type
            quiz.updated_by = user
            quiz.updated_at = timezone.now()
            quiz.total_marks = total_marks  # Update total marks if needed
            quiz.total_questions = 0  # Reset question count, will recalculate later
            if status is not None:
                if isinstance(status,bool):
                    quiz.status = status
                else:
                    return Response({"status": False, "message": "Invalid status value", "data": []})
            
            quiz.save()

            total_marks_accumulated = 0  
            total_questions = 0  

            if quiz_type == 'auto':
                # Handling 'auto' type quizzes
                if marks_breakdown:
                    # Clear existing questions if it's an "auto" quiz
                    old_questions = QuizQuestion.objects.filter(quiz=quiz)

                    for q in old_questions:
                        q.delete()

                    for marks, count in marks_breakdown.items():
                        marks = int(marks)  
                        count = int(count)  

                        questions = TrainingQuestions.objects.filter(
                            # training=training,  
                            # document=document,  
                            marks=marks,       
                            status=True,
                            document=quiz.document,
                        )

                        if len(questions) < count:
                            return Response({"status": False, "message": f"Not enough questions with {marks} marks. Found {len(questions)} questions.", "data": []})

                        questions_list = list(questions)
                        print(questions_list, "jjj")
                        selected_questions = random.sample(questions_list, count)
                        # selected_questions = questions_list

                        potential_marks = total_marks_accumulated + (marks * count)
                        if potential_marks > total_marks:
                            return Response({"status": False, "message": f"Total marks exceeded. The selected questions' marks total {potential_marks}, which exceeds the input total_marks of {total_marks}.", "data": []})

                        # For each selected question, update or add it
                        for question in selected_questions:
                            existing_question = old_questions.filter(question=question).first()
                            if existing_question:
                                # If the question already exists, do nothing, just update marks if needed
                                existing_question.marks = marks  # If you want to update marks, do it here
                                existing_question.save()
                            else:
                                # If the question is new, add it to the quiz
                                QuizQuestion.objects.create(quiz=quiz, question=question, marks=marks)
                            total_marks_accumulated += marks
                            total_questions += 1

            elif quiz_type == 'manual':
                # Handling 'manual' type quizzes
                if selected_questions:
                    questions = TrainingQuestions.objects.filter(
                        id__in=selected_questions,  
                        # training=training,  
                        # document=document,  
                        status=True
                    )

                    if len(questions) != len(selected_questions):
                        return Response({
                            "status": False,
                            "message": "Some of the selected questions are invalid or inactive.",
                            "data": []
                        })

                    # Add new or checked questions to the quiz
                    for question in questions:
                        if total_marks_accumulated + question.marks > total_marks:
                            return Response({
                                "status": False,
                                "message": f"Adding this question would exceed the total marks. Current total: {total_marks_accumulated}, question marks: {question.marks}",
                                "data": []
                            })
                        # Add the question if it's not already in the quiz
                        if not QuizQuestion.objects.filter(quiz=quiz, question=question).exists():
                            QuizQuestion.objects.create(quiz=quiz, question=question, marks=question.marks)
                            total_marks_accumulated += question.marks
                            total_questions += 1

                    # Remove unchecked questions
                    existing_quiz_questions = QuizQuestion.objects.filter(quiz=quiz)
                    for quiz_question in existing_quiz_questions:
                        # If a question is not selected anymore, remove it from the quiz
                        if quiz_question.question.id not in selected_questions:
                            quiz_question.delete()


            quiz.total_marks = total_marks_accumulated
            quiz.total_questions = total_questions
            quiz.save()

            serializer = TrainingQuizSerializer(quiz, context={'request': request})
            return Response({"status": True, "message": "Quiz updated successfully", "data": serializer.data})

        except TrainingQuiz.DoesNotExist:
            return Response({"status": False, "message": "Quiz not found", "data": []})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e), "data": []})

    def destroy(self, request, *args, **kwargs):
        try:
            training_quiz_id = self.kwargs.get("training_quiz_id")
            quiz = TrainingQuiz.objects.get(id=training_quiz_id)
            quiz.delete()
            return Response({"status": True, "message": "Quiz deleted successfully", "data": []})
        except TrainingQuiz.DoesNotExist:
            return Response({"status": False, "message": "Quiz not found", "data": []})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e), "data": []})
    
        
import pdfkit
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
from rest_framework.parsers import MultiPartParser
# from comtypes.client import CreateObject  # Windows-only alternative
         
class InductionCreateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = InductionSerializer
    queryset = Induction.objects.all().order_by('-id')

    def create(self, request, *args, **kwargs):
        try:
            induction_name = request.data.get('induction_name')
            department = request.data.get('department')
            document = request.FILES.get('document')

            department_instance = Department.objects.filter(id=department).first()
            if not department_instance:
                return Response({'status': False,'message': 'Invalid department'})

            if not induction_name:
                return Response({'status': False, 'message': 'Induction name is required'})
            if not department:
                return Response({'status': False,'message': 'Department is required'})
            if not document:
                return Response({'status': False,'message': 'Document is required'})
            

            # Create Induction
            induction = Induction.objects.create(
                department=department_instance,
                induction_name=induction_name,
                document=document,
            )

            serializer = InductionSerializer(induction, context={'request': request})
            return Response({"status": True, "message": "Induction created successfully", "data": serializer.data})

        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})

    def list(self, request):
        queryset = Induction.objects.all().order_by('-id')

        try:
            if queryset.exists():
                serializer = InductionSerializer(queryset, many=True,context={'request': request})
                return Response({
                    "status": True,
                    "message": "Induction fetched successfully",
                    "data": serializer.data
                })
            else:
                return Response({
                    "status": True,
                    "message": "No induction found",
                    "data": []
                })
        except Exception as e:
            return Response({"status": False, 'message': 'Something went wrong', 'error': str(e)})
class InductionUpdateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = InductionSerializer
    queryset = Induction.objects.all()
    lookup_field = 'id'

    def update(self, request, *args, **kwargs):
        try:
            induction = self.get_object()
            induction_name = request.data.get('induction_name')
            department = request.data.get('department')
            document = request.FILES.get('document')
    
            if induction_name:
                induction.induction_name = induction_name
    
            if department:  # Ensure department exists
                department_instance = Department.objects.filter(id=department).first()
                if department_instance:
                    induction.department = department_instance
                else:
                    return Response({'status': False, 'message': 'Invalid department ID'})
    
            if document:
                induction.document = document  # Handle document upload
    
            induction.save()
    
            serializer = InductionSerializer(induction, context={'request': request})
            return Response({"status": True, "message": "Induction updated successfully", "data": serializer.data})
    
        except Exception as e:
            return Response({"status": False, "message": f"Something went wrong: {str(e)}", "data": []})


    def destroy(self, request, *args, **kwargs):
        try:
            induction = self.get_object()
            induction.delete()
            return Response({"status": True, "message": "Induction deleted successfully"})
        except Exception as e:
            return Response({"status": False, "message": f"Something went wrong: {str(e)}"})
class InductionDesignationCreateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = InductionDesignationSerializer
    queryset = InductionDesignation.objects.all().order_by('-id')

    def create(self, request, *args, **kwargs):
        try:
            induction_designation_name = request.data.get('induction_designation_name')
            designation_code = request.data.get('designation_code')
            induction_id = request.data.get('induction')

            # Validation
            if not induction_designation_name:
                return Response({'status': False, 'message': 'Induction Designation Name is required'})
            if not designation_code:
                return Response({'status': False, 'message': 'Designation Code is required'})
            if not induction_id:
                return Response({'status': False, 'message': 'Induction ID is required'})

            induction_designation = InductionDesignation.objects.create(
                induction_designation_name=induction_designation_name,
                designation_code=designation_code,
                induction_id=induction_id,
                created_by=request.user
            )

            serializer = self.get_serializer(induction_designation)
            return Response({"status": True, "message": "Induction Designation created successfully", "data": serializer.data})

        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        
    def list(self, request):
        queryset = InductionDesignation.objects.all().order_by('-id')
        
        try:
            if queryset.exists():
                serializer = InductionDesignationSerializer(queryset, many=True)
                return Response({
                    "status": True,
                    "message": "Induction designations fetched successfully",
                    "total": queryset.count(),
                    "data": serializer.data
                })
            else:
                return Response({
                    "status": True,
                    "message": "No induction designation found",
                    "total": 0,
                    "data": []
                })
        except Exception as e:
            return Response({"status": False, 'message': 'Something went wrong', 'error': str(e)})      

class InductionDesignationUpdateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = InductionDesignationSerializer
    queryset = InductionDesignation.objects.all()
    lookup_field = 'id'

    def update(self, request, *args, **kwargs):
        try:
            induction_designation = self.get_object()
            induction_designation_name = request.data.get('induction_designation_name')
            designation_code = request.data.get('designation_code')
            induction_id = request.data.get('induction')

            if induction_designation_name:
                induction_designation.induction_designation_name = induction_designation_name
            if designation_code:
                induction_designation.designation_code = designation_code
            if induction_id:
                induction_designation.induction_id = induction_id

            induction_designation.save()
            serializer = self.get_serializer(induction_designation)
            return Response({"status": True, "message": "Induction Designation updated successfully", "data": serializer.data})

        except Exception as e:
            return Response({"status": False, "message": f"Something went wrong: {str(e)}"})

    def destroy(self, request, *args, **kwargs):
        # Check if the user has permission to delete InductionDesignations
        # if not request.user.has_perm('your_app.delete_inductiondesignation'):
        #     return Response({"status": False, "message": "You are not authorized to delete this designation!"}, status=status.HTTP_403_FORBIDDEN)
        
        try:
            # Get the InductionDesignation ID from the URL
            induction_id = self.kwargs.get("id")
            
            # Check if the InductionDesignation exists
            if not InductionDesignation.objects.filter(id=induction_id).exists():
                return Response({"status": False, "message": "InductionDesignation ID not found"})

            # Delete the InductionDesignation
            InductionDesignation.objects.filter(id=induction_id).delete()
            return Response({"status": True, "message": "InductionDesignation deleted successfully"})
        
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})       


import ipdb

class ClassroomCreateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ClassroomTrainingSerializer

    def create(self, request, *args, **kwargs):
        try:
            user = request.user
            classroom_name = request.data.get('classroom_name')
            document_id = request.data.get('document_id')
            is_assesment = request.data.get('is_assesment')
            description = request.data.get('description')
            upload_doc = request.FILES.getlist('upload_doc')
            trainer_id = request.data.get('trainer')
            online_offline_status = request.data.get('online_offline_status')
            status = request.data.get('status')
            select_users = request.data.get('select_users', [])

            # Validate required fields
            if not classroom_name:
                return Response({'status': False, 'message': 'Classroom name is required.'})
            if not is_assesment:
                return Response({'status': False, 'message': 'Is assessment is required.'})
            if not description:
                return Response({'status': False, 'message': 'Description is required.'})
            if not trainer_id:
                return Response({'status': False, 'message': 'Trainer ID is required.'})
            if not online_offline_status:
                return Response({'status': False, 'message': 'Online/offline status is required.'})
            if online_offline_status == 'offline' and not upload_doc:
                return Response({'status': False, 'message': 'Please upload document for offline training.'})
            if not select_users:
                return Response({'status': False, 'message': 'Please select users for No document training.'})

            # Convert select_users if it's a string
            if isinstance(select_users, str):
                try:
                    select_users = json.loads(select_users)
                except json.JSONDecodeError:
                    return Response({'status': False, 'message': 'Invalid format for select_users. Expected a list of user IDs.'})

            # Validate trainer
            trainer = Trainer.objects.filter(id=trainer_id).first()
            if not trainer:
                return Response({'status': False, 'message': 'Invalid trainer selected.'})

            # Handle document_id correctly
            if document_id in [None, "", "None"]:
                document_id = None
            else:
                try:
                    document_id = int(document_id)
                except ValueError:
                    return Response({'status': False, 'message': 'Invalid document ID format. Expected an integer or None.'})

            # Fetch document if document_id is provided
            document = None
            if document_id is not None:
                document = Document.objects.filter(id=document_id).first()
                if not document:
                    return Response({'status': False, 'message': 'Invalid document selected.'})

            # Fetch users
            users = CustomUser.objects.filter(id__in=select_users)
            if document_id is None and not users.exists():
                return Response({'status': False, 'message': 'Invalid users selected.'})

            # Create ClassroomTraining instance
            classroom_training = ClassroomTraining.objects.create(
                classroom_name=classroom_name,
                is_assesment=is_assesment,
                description=description,
                status=status,
                trainer=trainer,
                document=document,
                online_offline_status=online_offline_status,
            )

            # Assign users to classroom
            if select_users:
                classroom_training.user.set(users)

            # Save uploaded documents
            for file in upload_doc:
                ClassroomTrainingFile.objects.create(classroom_training=classroom_training, upload_doc=file)

            serializer = ClassroomTrainingSerializer(classroom_training, context={'request': request})
            return Response({"status": True, "message": "Classroom training created successfully", "data": serializer.data})

        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        
    def list(self, request):
        user = request.user
        if user.groups.filter(name="DTC").exists():
            queryset = ClassroomTraining.objects.all().order_by('-id')
        else:
            queryset = ClassroomTraining.objects.filter(user=user).order_by('-id')
        
        try:
            if queryset.exists():
                serializer = ClassroomTrainingSerializer(queryset, many=True, context={'request': request})
                return Response({"status": True,"message": "Classroom training fetched successfully","data": serializer.data})
            else:
                return Response({"status": True,"message": "No classroom training found","data": []})
            
        except Exception as e:
            return Response({"status": False, 'message': 'Something went wrong', 'error': str(e)})
        
class ClassroomUpdateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ClassroomTrainingSerializer
    queryset = ClassroomTraining.objects.all()
    lookup_field = 'id'

    def update(self, request, *args, **kwargs):
        try:
            user = request.user
            if not user.groups.filter(name="DTC").exists():
                return Response({"status": False, "message": "You do not have permission to view these documents."})
            classroom_training = self.get_object()
            classroom_name = request.data.get('classroom_name')
            is_assesment = request.data.get('is_assesment')
            description = request.data.get('description')
            upload_doc = request.FILES.getlist('upload_doc')
            status = request.data.get('status')
            if classroom_name:
                classroom_training.classroom_name = classroom_name
            if is_assesment:
                classroom_training.is_assesment = is_assesment
            if description:
                classroom_training.description = description
            if status:
                classroom_training.status = status

            classroom_training.save()

            if upload_doc:
                for file in upload_doc:
                    ClassroomTrainingFile.objects.create(classroom_training=classroom_training, upload_doc=file)

            serializer = ClassroomTrainingSerializer(classroom_training, context={'request': request})
            return Response({"status": True, "message": "Classroom training updated successfully", "data": serializer.data})

        except Exception as e:
            return Response({"status": False, "message": f"Something went wrong: {str(e)}"})
        
    def destroy(self, request, *args, **kwargs):
        try:
            classroom_training = self.get_object()
            classroom_training.delete()
            return Response({"status": True, "message": "Classroom deleted successfully"})
        except Exception as e:
            return Response({"status": False, "message": f"Something went wrong: {str(e)}"})

    # def create(self, request, *args, **kwargs):
    #     try:
    #         title = request.data.get('title')
    #         department_or_employee = request.data.get('department_or_employee')
    #         training_type = request.data.get('classroom_training_type')
    #         description = request.data.get('description')
    #         sop = request.data.get('sop')
    #         start_date = request.data.get('start_date')
    #         start_time = request.data.get('start_time')
    #         end_time = request.data.get('end_time')
    #         document = request.data.get('document')
    #         status = request.data.get('status', 'assigned')
    #         created_by = request.user.id

    #         # Validation
    #         if not title or not department_or_employee or not training_type or not description:
    #             return Response({'status': False, 'message': 'All fields are required.'})

    #         # Create Classroom Training
    #         classroom_training = ClassroomTraining.objects.create(
    #             title=title,
    #             department_or_employee_id=department_or_employee,
    #             classroom_training_type=training_type,
    #             description=description,
    #             sop=sop,
    #             start_date=start_date,
    #             start_time=start_time,
    #             end_time=end_time,
    #             created_by_id=created_by,
    #             status=status
    #         )

    #         if document:
    #             classroom_training.document = document
    #             classroom_training.save()

    #         serializer = ClassroomTrainingSerializer(classroom_training, context={'request': request})
    #         return Response({
    #             "status": True,
    #             "message": "Classroom training created successfully",
    #             "data": serializer.data
    #         })

    #     except Exception as e:
    #         return Response({"status": False, "message": "Something went wrong", "error": str(e)})


    # def mark_completed(self, request, *args, **kwargs):
    #     try:
    #         # Get the classroom training object
    #         classroom_training = self.get_object()
    #         if classroom_training.classroom_training_type == "assessment":
    #             # Check if all users have provided assessment results
    #             users = classroom_training.department_or_employee.users.all()  # Assuming `users` is a related field
    #             missing_assessment = []

    #             for user in users:
    #                 if not user.assessment_result:  # Assuming `assessment_result` is the field on the user model
    #                     missing_assessment.append(user.username)

    #             if missing_assessment:
    #                 return Response({
    #                     "status": False,
    #                     "message": f"Please provide assessment results for the following users: {', '.join(missing_assessment)}"
    #                 })

    #         # If no missing assessment or no assessment type, change status to completed
    #         classroom_training.status = 'completed'
    #         classroom_training.save()

    #         return Response({
    #             "status": True,
    #             "message": "Classroom training status updated to completed successfully"
    #         })
    #     except Exception as e:
    #         return Response({
    #             "status": False,
    #             "message": f"Something went wrong: {str(e)}"
    #         })

class SessionCreateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = SessionSerializer
    queryset = Session.objects.all().order_by('-id')

    def create(self, request, *args, **kwargs):
        try:
            user = request.user
            if not user.groups.filter(name="DTC").exists():
                return Response({"status": False, "message": "You do not have permission to view these documents."})
            session_name = request.data.get('session_name')
            venue = request.data.get('venue')
            start_date = request.data.get('start_date')
            # end_date = request.data.get('end_date')
            start_time = request.data.get('start_time')
            # end_time = request.data.get('end_time')
            user_ids = request.data.get('user_ids')
            classroom_id = request.data.get('classroom_id')
            print(user_ids)
            print(f"user_ids type: {type(user_ids)}")
            if not session_name:
                return Response({'status': False, 'message': 'Session name is required.'})
            if not venue:
                return Response({'status': False, 'message': 'Venue is required.'})
            if not user_ids or not isinstance(user_ids, list):
                return Response({'status': False, 'message': 'User is required.'})
            if not classroom_id:
                return Response({'status': False, 'message': 'Classroom ID is required.'})

            try:
                classroom = ClassroomTraining.objects.get(id=classroom_id)
            except ClassroomTraining.DoesNotExist:
                return Response({"status": False, "message": "Classroom not found."})
            
            session = Session.objects.create(
                session_name=session_name,
                venue=venue,
                start_date=start_date,
                # end_date=end_date,
                start_time=start_time,
                # end_time=end_time,
                classroom_id=classroom_id
            )
            users = CustomUser.objects.filter(id__in=user_ids)
            session.user_ids.set(users)
            session.save()
            if not users.exists():
                return Response({"status": False, "message": "One or more user IDs are invalid."})
            
            session.save()

            serializer = SessionSerializer(session, context={'request': request})
            return Response({"status": True, "message": "Session created successfully", "data": serializer.data})

        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})

    def list(self, request, *args, **kwargs):
        try:
            classroom_id = request.query_params.get('classroom_id')

            if classroom_id:
                queryset = self.queryset.filter(classroom_id=classroom_id)
            else:
                return Response({"status": False, "message": "classroom_id is required."})
            
            session_data = []
            for session in queryset:
                session_info = {
                    "session_id": session.id,
                    "session_name": session.session_name,
                    "venue": session.venue,
                    "start_date": session.start_date,
                    "start_time": session.start_time,
                    "attend": session.attend
                }
                user_ids = session.user_ids.values_list('id', flat=True)
                session_info["user_ids"] = list(user_ids)

                user = request.user
                session_complete = SessionComplete.objects.filter(session=session).first()
                session_info["is_completed"] = session_complete.is_completed if session_complete else False
                
                session_data.append(session_info)

            if session_data:
                return Response({"status": True, "message": "Sessions fetched successfully", "data": session_data})
            else:
                return Response({"status": True, "message": "No sessions found", "data": []})

        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        
class SessionCompletedViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = SessionCompleteSerializer
    queryset = SessionComplete.objects.all().order_by('-id')

    def mark_completed(self, request, *args, **kwargs):
        try:
            user = request.user
            if not user.groups.filter(name="DTC").exists():
                return Response({"status": False, "message": "You do not have permission to view these documents."})
            session_id = self.kwargs.get('session_id')
            session_instance = Session.objects.filter(id=session_id).first()
            if not session_instance:
                return Response({"status": False, "message": "Session not found."})
            
            session_complete = SessionComplete.objects.create(session=session_instance, is_completed=True)

            serializer = SessionCompleteSerializer(session_complete, context={'request': request})
            return Response({"status": True, "message": "Session completed successfully", "data": serializer.data})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
    


class SessionUpdateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = SessionSerializer
    queryset = Session.objects.all()
    lookup_field = 'id'

    def update(self, request, *args, **kwargs):
        try:
            user = request.user
            if not user.groups.filter(name="DTC").exists():
                return Response({"status": False, "message": "You do not have permission to view these documents."})
            session = self.get_object()
            session_name = request.data.get('session_name')
            venue = request.data.get('venue')
            start_date = request.data.get('start_date')
            end_date = request.data.get('end_date')
            start_time = request.data.get('start_time')
            end_time = request.data.get('end_time')
            user_ids = request.data.get('user_ids')
            classroom_id = request.data.get('classroom_id')

            if session_name:
                session.session_name = session_name
            if venue:
                session.venue = venue
            if start_date:
                session.start_date = start_date
            if end_date:
                session.end_date = end_date
            if start_time:
                session.start_time = start_time
            if end_time:
                session.end_time = end_time
            if classroom_id:
                session.classroom_id = classroom_id

            session.save()

            if user_ids:
                users = CustomUser.objects.filter(id__in=user_ids)
                session.user_ids.set(users)
                if not users.exists():
                    return Response({"status": False, "message": "One or more user IDs are invalid."})

            serializer = SessionSerializer(session, context={'request': request})
            return Response({"status": True, "message": "Session updated successfully", "data": serializer.data})

        except Exception as e:
            return Response({"status": False, "message": f"Something went wrong: {str(e)}"})

    def destroy(self, request, *args, **kwargs):
        try:
            session = self.get_object()
            session.delete()
            return Response({"status": True, "message": "Session deleted successfully"})
        except Exception as e:
            return Response({"status": False, "message": f"Something went wrong: {str(e)}"})

class AttendanceCreateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def create(self, request, *args, **kwargs):
        try:
            session_id = request.data.get('session_id')
            user_ids = request.data.get('user_ids')
            status = request.data.get('status')
            print(f"user_ids type: {type(user_ids)}")
            if not session_id or not user_ids:
                return Response({'status': False, 'message': 'Session ID and user IDs are required.'})

            if status not in ['present', 'absent']:
                return Response({'status': False, 'message': 'Invalid status value.'})
            
            if isinstance(user_ids, str):
                try:
                    user_ids = [int(uid) for uid in user_ids.split(',')]
                except ValueError:
                    return Response({'status': False, 'message': 'Invalid user_ids format.'})

            if not isinstance(user_ids, list) or not all(isinstance(uid, int) for uid in user_ids):
                return Response({'status': False, 'message': 'user_ids should be a list of integers.'})
            
            try:
                session = Session.objects.get(id=session_id)
            except Session.DoesNotExist:
                return Response({"status": False, "message": "Session not found."})
            
            users = CustomUser.objects.filter(id__in=user_ids)
            if not users.exists():
                return Response({"status": False, "message": "One or more user IDs are invalid."})
            
            for user in users:
                attendance, created = Attendance.objects.get_or_create(user=user, session=session)
                # quiz_attempts, created = QuizSession.objects.get_or_create(user=user, quiz=quiz)
                attendance.status = status
                attendance.save()

            classroom_training = ClassroomTraining.objects.filter(user__in=users, sessions__id=session_id).distinct().first()
            classroom_training.is_all_completed = True
            classroom_training.save()
            if classroom_training:
                all_present = Attendance.objects.filter(session=session, status='absent').count() == 0
                

            if Attendance.objects.filter(session=session, status='present').exists():
                session.attend = True
            else:
                session.attend = False
            session.save()

            return Response({"status": True, "message": "Attendance marked successfully."})
        
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})

    def list(self, request, *args, **kwargs):
        try:
            session_id = request.query_params.get('session_id')

            if session_id:
                session = Session.objects.get(id=session_id)
                all_users = session.user_ids.all()

                attendance_data = []
                for user in all_users:
                    attendance = Attendance.objects.filter(session=session, user=user).first()
                    attendance_data.append({
                        "user_id": user.id,
                        "user_name": user.username,
                        "status": attendance.status if attendance else "Absent",
                    })
            else:
                return Response({"status": False, "message": "session_id is required."})
            
            return Response({"status": True, "message": "Attendance fetched successfully", "data": attendance_data})
        
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        
# class ClassroomTrainingUpdateViewSet(viewsets.ModelViewSet):
#     permission_classes = [permissions.IsAuthenticated]
#     serializer_class = ClassroomTrainingSerializer
#     queryset = ClassroomTraining.objects.all()
#     lookup_field = 'id'

#     def update(self, request, *args, **kwargs):
#         try:
#             classroom_training = self.get_object()
#             training_type = request.data.get('classroom_training_type')
#             title = request.data.get('title')
#             department_or_employee = request.data.get('department_or_employee')
#             description = request.data.get('description')
#             sop = request.data.get('sop')
#             start_date = request.data.get('start_date')
#             start_time = request.data.get('start_time')
#             end_time = request.data.get('end_time')
#             status = request.data.get('status')
#             acknowledged_by_employee = request.data.get('acknowledged_by_employee')

#             # Updating fields
#             if training_type: classroom_training.classroom_training_type = training_type
#             if title: classroom_training.title = title
#             if department_or_employee: classroom_training.department_or_employee_id = department_or_employee
#             if description: classroom_training.description = description
#             if sop: classroom_training.sop = sop
#             if start_date: classroom_training.start_date = start_date
#             if start_time: classroom_training.start_time = start_time
#             if end_time: classroom_training.end_time = end_time
#             if status: classroom_training.status = status
#             if acknowledged_by_employee is not None:
#                 classroom_training.acknowledged_by_employee = acknowledged_by_employee

#             classroom_training.save()
#             serializer = ClassroomTrainingSerializer(classroom_training, context={'request': request})
#             return Response({"status": True, "message": "Classroom training updated successfully", "data": serializer.data})

#         except Exception as e:
#             return Response({"status": False, "message": f"Something went wrong: {str(e)}", "data": []})

#     def destroy(self, request, *args, **kwargs):
#         try:
#             classroom_training = self.get_object()
#             classroom_training.delete()
#             return Response({"status": True, "message": "Classroom training deleted successfully"})
#         except Exception as e:
#             return Response({"status": False, "message": f"Something went wrong: {str(e)}"})
        
# class TrainingListViewSet(viewsets.ModelViewSet):
#     permission_classes = [permissions.IsAuthenticated]
#     serializer_class = TrainingListSerializer
#     queryset = TrainingCreate.objects.all()

#     def list(self, request, *args, **kwargs):
#         plant_id = request.data.get('plant')
#         training_type_id = request.data.get('type')
#         training_number = request.data.get('training_number')

#         filters = Q()
#         if plant_id:
#             filters &= Q(plant_id=plant_id)
#         if training_type_id:
#             filters &= Q(training_type_id=training_type_id)
#         if training_number:
#             filters &= Q(training_number=training_number)

#         trainings = self.queryset.filter(filters)

#         serializer = self.get_serializer(trainings, many=True)
#         return Response({
#             "status": True,
#             "message": "Training data fetched successfully",
#             "data": serializer.data
#         })
        
class TrainingListViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingListSerializer
    queryset = TrainingCreate.objects.all()

    def list(self, request, *args, **kwargs):
        trainings = self.queryset  # No filtering logic
        serializer = self.get_serializer(trainings, many=True)
        return Response({
            "status": True,
            "message": "Training data fetched successfully",
            "data": serializer.data
        })


class TrainingMatrixAssignUserViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    queryset = TrainingMatrix.objects.all()
    serializer_class = TrainingMatrixAssignUserSerializer

    def create(self, request, *args, **kwargs):
        user = self.request.user
        training_create_id = request.data.get('training_create_id')
        user_ids = request.data.get('user_ids', [])
        training_duration = request.data.get('training_duration', None)
        evaluation_status = request.data.get('evaluation_status', None)
        assigned_role_id = request.data.get('assigned_role', None)
        due_reason = request.data.get('due_reason', None)

        # Validate the training_create_id
        if not training_create_id:
            return Response({"status": False, "message": "TrainingCreate ID not provided.", "data": []})

        try:
            training_create_instance = TrainingCreate.objects.get(id=training_create_id)
        except TrainingCreate.DoesNotExist:
            return Response({
                "status": False, "message": "TrainingCreate instance not found.", "data": []})

        # Validate user_ids
        if not user_ids:
            return Response({"status": False, "message": "User IDs not provided.", "data": []})

        users = CustomUser.objects.filter(id__in=user_ids)
        if users.count() != len(user_ids):
            return Response({"status": False, "message": "Invalid user IDs.", "data": []})

        # Validate assigned_role_id if provided
        assigned_role = None
        if assigned_role_id:
            try:
                assigned_role = JobRole.objects.get(id=assigned_role_id)
            except JobRole.DoesNotExist:
                return Response({
                    "status": False, "message": "Assigned role not found.", "data": []})

        # Check if the TrainingMatrix already exists
        training_matrix = TrainingMatrix.objects.filter(training=training_create_instance).first()

        if training_matrix:
            # Update existing TrainingMatrix with new data
            training_matrix.assigned_user.set(users)  # Update the assigned users
            training_matrix.training_duration = training_duration
            training_matrix.evaluation_status = evaluation_status
            training_matrix.assigned_role = assigned_role
            training_matrix.due_reason = due_reason
            training_matrix.save()

            message = "Training matrix updated successfully."
        else:
            # Create new TrainingMatrix if not exists
            training_matrix = TrainingMatrix.objects.create(
                training=training_create_instance,
                assigned_by=user,
                training_duration=training_duration,
                evaluation_status=evaluation_status,
                assigned_role=assigned_role,
                due_reason=due_reason
            )
            training_matrix.assigned_user.set(users)
            training_matrix.save()

            message = "Training matrix created or users assigned successfully."

        # Serialize and return the response
        serializer = TrainingMatrixAssignUserSerializer(training_matrix)
        return Response({"status": True, "message": message, "data": serializer.data})


    def list(self, request, *args, **kwargs):
        training_matrix = TrainingMatrix.objects.all()
        serializer = TrainingMatrixAssignUserSerializer(training_matrix, many=True)
        data = serializer.data
        return Response({"status": True, "message": "Training matrix data fetched successfully", "data": data})
    

class TrainingWiseTrainingMatrixViewset(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingMatrixAssignUserSerializer

    def list(self, request, *args, **kwargs):
        try:
            training_id = request.query_params.get('training_id')
            if not training_id:
                return Response({"status": False, "message": "Training ID not provided.", "data": []})

            training_matrix = TrainingMatrix.objects.filter(training_id=training_id)
            serializer = TrainingMatrixAssignUserSerializer(training_matrix, many=True)
            data = serializer.data

            return Response({"status": True, "message": "Training matrix data fetched successfully", "data": data})
        except Exception as e:
            return Response({"status": False, "message": str(e), "data": []})
    

class JobroleListingViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = GetJobRoleSerializer

    def list(self, request, *args, **kwargs):
        plant_id = request.data.get('plant')
        department_id = request.data.get('department')
        area_id = request.data.get('area')

        job_roles = JobRole.objects.all()

        if plant_id:
            job_roles = job_roles.filter(plant_id=plant_id)
        if department_id:
            job_roles = job_roles.filter(department_id=department_id)
        if area_id:
            job_roles = job_roles.filter(area_id=area_id)

        job_role_serializer = GetJobRoleSerializer(job_roles, many=True)

        return Response({
            "status": True,
            "message": "Training and job role data fetched successfully",
            "data": {
                "job_roles": job_role_serializer.data
                }
        })



class TrainingAssignViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'user_id'
    def update(self, request, *args, **kwargs):
        try:
            user_id = self.kwargs.get("user_id")
            if not user_id:
                return Response({"status": False, "message": "user_id is missing"})
            
            user_instance = CustomUser.objects.filter(id=user_id).first()
            if not user_instance:
                return Response({"status": False, "message": "user_id not found"})
            job_assign_instance, created = JobAssign.objects.get_or_create(user=user_instance)
            
            job_role_ids = request.data.get('job_role_ids', [])
            # remove_job_role_ids = request.data.get('remove_job_role_ids', [])
            if not isinstance(job_role_ids, list):
                return Response({"status": False, "message": "Job role IDs should be a list"})

            valid_job_roles = JobRole.objects.filter(id__in=job_role_ids)
            # remove_job_role_ids = JobRole.objects.filter(id__in=remove_job_role_ids)

            if len(valid_job_roles) != len(job_role_ids):
                return Response({
                    "status": False,
                    "message": "Some Job Role IDs are invalid"
                })

            job_assign_instance.job_roles.clear()
            job_assign_instance.job_roles.add(*valid_job_roles)
            # job_assign_instance.job_roles.remove(*remove_job_role_ids)
            user_instance.is_jr_assign = True
            user_instance.save()
            job_assign_instance.save()

            return Response({
                "status": True,
                "message": "Job roles updated for the user successfully",
            })
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})


class TrainingAssignListViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = JobAssignSerializer
    queryset = JobAssign.objects.all()

    def list(self, request, *args, **kwargs):
        try:
            user_id = self.kwargs.get("user_id")
            if not user_id:
                return Response({"status": False,"message": "user_id is missing"})
            
            user_instance = CustomUser.objects.filter(id=user_id).first()
            if not user_instance:
                return Response({"status": False,"message": "user_id not found"})
            
            queryset = self.queryset.filter(user=user_instance)
            
            
            job_assign_serializer = self.serializer_class(queryset, many=True, context={'request': request})
            
            return Response({"status": True,"message": "Job roles assigned to the user fetched successfully","data": job_assign_serializer.data})
        except Exception as e:
            return Response({"status": False,"message": str(e),"data": []})
    

# class JobroleListingapiViewSet(viewsets.ModelViewSet):
#     permission_classes = [permissions.IsAuthenticated]
#     serializer_class = GetJobRoleSerializer

#     def list(self, request, *args, **kwargs):
#         plant_id = request.data.get('plant')
#         department_id = request.data.get('department')
#         area_id = request.data.get('area')
#         job_role_name = request.data.get('job_role_name')

#         job_roles = JobRole.objects.all()

#         if plant_id:
#             job_roles = job_roles.filter(plant_id=plant_id)
#         if department_id:
#             job_roles = job_roles.filter(department_id=department_id)
#         if area_id:
#             job_roles = job_roles.filter(area_id=area_id)
#         if job_role_name:
#             job_roles = job_roles.filter(job_role_name__icontains=job_role_name)

#         job_role_serializer = self.serializer_class(job_roles, many=True)

#         return Response({
#             "status": True,
#             "message": "Training and job role data fetched successfully",
#             "data": {
#                 "job_roles": job_role_serializer.data
#             }
#         })

class JobroleListingapiViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingdataSerializer

    def list(self, request, *args, **kwargs):
        job_role_id = request.data.get('job_role_id')
        
        if not job_role_id:
            return Response({
                "status": False,
                "message": "Job role ID is required",
                "data": []
            })

        # Filter TrainingCreate objects by job_role_id
        trainings = TrainingCreate.objects.filter(job_roles__id=job_role_id).distinct()

        # Serialize the data
        serializer = self.serializer_class(trainings, many=True)

        return Response({
            "status": True,
            "message": "Trainings fetched successfully",
            "data": serializer.data
        })


class TrainingListingViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TrainingSerializer

    def list(self, request, *args, **kwargs):
        training_type_id = request.data.get("training_type")

        trainings = TrainingCreate.objects.all()

        if training_type_id:
            trainings = trainings.filter(training_type_id=training_type_id)

        # Serialize the data
        training_serializer = self.serializer_class(trainings, many=True)

        # Return the response
        return Response({
            "status": True,
            "message": "Training data fetched successfully",
            "data": training_serializer.data
        })

class TrainingAssigntoJobroleViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    queryset = TrainingCreate.objects.all()
    serializer_class = TrainingCreateSerializer

    def create(self, request, *args, **kwargs):

        try:
            job_role_id = request.data.get('job_role_id')
            if not job_role_id:
                return Response({"status": False, "message": "Job role ID is required"})

            job_role_instance = JobRole.objects.filter(id=job_role_id).first()
            if not job_role_instance:
                return Response({"status": False, "message": "Invalid Job role ID"})

            document_ids = request.data.get('document_ids', [])
            if not isinstance(document_ids, list) or not document_ids:
                return Response({"status": False, "message": "document_ids should be a non-empty list"})

            existing_documents = Document.objects.filter(job_roles=job_role_instance)
            for doc in existing_documents:
                doc.job_roles.remove(job_role_instance)
                doc.is_effective = False
                doc.save()

            valid_document = Document.objects.filter(id__in=document_ids)
            if len(valid_document) != len(document_ids):
                return Response({"status": False, "message": "Some Training IDs are invalid"})

            for training in valid_document:
                training.job_roles.add(job_role_instance)
                training.is_effective = True
                training.save()

            return Response({
                "status": True,
                "message": "document successfully assigned to the job role",
            })
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        
    def list(self, request, *args, **kwargs):
        try:
            job_role_id = request.query_params.get('job_role_id')
            if not job_role_id:
                return Response({"status": False, "message": "Job role ID is required"})

            job_role_instance = JobRole.objects.filter(id=job_role_id).first()
            if not job_role_instance:
                return Response({"status": False, "message": "Invalid Job role ID"})

            assigned_documents = Document.objects.filter(job_roles=job_role_instance)
            # serialized_documents = DocumentSerializer(assigned_documents, many=True)
            serialized_documents = DocumentMappingSerializer(assigned_documents, many=True)
            

            return Response({
                "status": True,
                "message": "Documents retrieved successfully",
                "documents": serialized_documents.data
            })
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})

class DocumentHasQuizListViewSet(viewsets.ModelViewSet):
    # permission_classes = [permissions.IsAuthenticated]
    serializer_class = DocumentMappingSerializer
    def list(self, request, *args, **kwargs):
        try:
            training = Document.objects.filter(trainingquiz__isnull=False).exclude(document_current_status=15).exclude(document_current_status=12).distinct()
            serializer = self.serializer_class(training, many=True)
            return Response({"status": True, "message": "Training with quizzes fetched successfully", "document_data": {"documents": serializer.data}})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})


class MaterialStartStopReadingView(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, *args, **kwargs):
        material_id = request.data.get('material_id')
        action = request.data.get('action')  # "start_reading" or "stop_reading"
        start_time = request.data.get('start_time', None)
        end_time = request.data.get('end_time', None)
        # Fetch the material and user
        user = self.request.user
        try:
            material = TrainingMaterial.objects.get(id=material_id)
        except TrainingMaterial.DoesNotExist:
            return Response({"status": False, "message": "Material not found.", "data": []})
        
        # Check if the material is part of the current training
        if material not in material.section.all():
            return Response({"status": False, "message": "Material is not part of the selected training section.", "data": []})

        if action == "start_reading":
            # Start reading the material
            session = MaterialReadingTime.objects.create(
                user=user,
                material=material,
                start_time=start_time
            )
            return Response({
                "status": True,
                "message": "Started reading the material.",
                "data": {"session_id": session.id, "material": material.material_title}
            })

        elif action == "stop_reading":
            # Stop reading the material
            session = MaterialReadingTime.objects.filter(user=user, material=material, end_time__isnull=True).first()
            if not session:
                return Response({"status": False, "message": "Reading session not found."}, status=404)

            session.end_time = end_time
            session.time_spent = session.end_time - session.start_time
            session.save()

            return Response({
                "status": True,
                "message": "Stopped reading the material.",
                "data": {"session_id": session.id, "material": material.material_title, "time_spent": session.time_spent}
            })
    

class TrainingStatusUpdateViewset(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    queryset = TrainingCreate.objects.all()
    lookup_field = 'training_id'

    def update(self, request, *args, **kwargs):
        training_id = self.kwargs.get('training_id')
        action = request.data.get('action')
        start_time = request.data.get('start_time')
        end_time = request.data.get('end_time')

        # Fetch the training instance
        try:
            training = TrainingCreate.objects.get(id=training_id)
        except TrainingCreate.DoesNotExist:
            return Response({"status": False, "message": "Training not found.","data":[]})

        if action == "start_training":
            if training.training_status == "in_progress":
                return Response({"status": False, "message": "Training already in progress.","data":[]})

            training.training_status = "in_progress"
            training.start_time = start_time
            training.save()

            return Response({
                "status": True,
                "message": "Training started.",
                "data": {"training_id": training.id, "training_status": training.training_status}
            })

        elif action == "stop_training":
            if training.training_status == "Completed":
                return Response({"status": False, "message": "Training already completed.","data": {"training_id": training.id, "training_status": training.training_status}}, status=400)

            # Check if all materials are read
            all_materials_read = not MaterialReadingTime.objects.filter(material__section__training=training, end_time__isnull=True).exists()
            if all_materials_read:
                training.training_status = "Completed"
                training.end_time = end_time
                training.save()

                return Response({
                    "status": True,
                    "message": "Training completed.",
                    "data": {"training_id": training.id, "training_status": training.training_status}
                })
            else:
                return Response({
                    "status": False,
                    "message": "Not all materials have been read yet.",
                    "data":[]
                })

        return Response({"status": False, "message": "Invalid action.","data": []})




class StartExam(viewsets.ModelViewSet):
    queryset = QuizSession.objects.all()
    serializer_class = QuizSessionSerializer

    def create(self, request, *args, **kwargs):
        quiz_id = request.data.get('quiz_id')
        quiz = TrainingQuiz.objects.get(id=quiz_id)
        user = self.request.user
        assigned_document = quiz.document
        assigned_document_version = assigned_document.version

        previous_session = QuizSession.objects.filter(user=user, quiz=quiz).order_by('-id').first()
        if previous_session:
            previous_major_version = previous_session.document_version.split('.')[0]
            current_major_version = assigned_document_version.split('.')[0]
            if previous_major_version != current_major_version:
                user.is_tni_consent = True
                user.save()
                return Response({"status": True,"message": "Exam started successfully.","quiz_session_id": quiz_session.id})
            
        attempts_count = QuizSession.objects.filter(user=user, quiz=quiz).count()

        if attempts_count < 3:
            quiz_session = QuizSession.objects.create(user=user, quiz=quiz, attempts=attempts_count + 1)
            user.is_tni_consent = True
            user.save()
            return Response({
                "status": True,
                "message": "Exam started successfully.",
                "quiz_session_id": quiz_session.id
            })

        session_complete = SessionComplete.objects.filter(session__quiz=quiz, user=user, is_completed=True).first()
        if not session_complete:
            return Response({
                "status": False,
                "message": "You must complete the session before starting the exam again."})

        quiz_session = QuizSession.objects.create(user=user, quiz=quiz, attempts=attempts_count + 1)
        user.is_tni_consent = True
        user.save()
        return Response({"status": True,"message": "Exam started successfully.","quiz_session_id": quiz_session.id})
    

    def list(self, request, *args, **kwargs):
        session_id = self.kwargs.get('session_id')

        quiz_session = QuizSession.objects.get(id=session_id)
        data = TrainingQuizSerializer(quiz_session.quiz).data


        return Response({
            
        })




class GetNextQuestion(viewsets.ModelViewSet):
    queryset = QuizSession.objects.all()
    serializer_class = QuizSessionSerializer

    def list(self, request, *args, **kwargs):
        # Get the QuizSession instance
        session_id = self.kwargs.get('session_id')
        quiz_session = QuizSession.objects.get(id=session_id)
        quiz = quiz_session.quiz

        # Get the questions for this quiz and fetch the next question
        questions = quiz.questions.all()  # Using related_name `questions` on the Quiz model

        # Check if there are more questions to show
        if quiz_session.current_question_index < len(questions):
            current_question = questions[quiz_session.current_question_index]

            # Prepare the data for the next question
            question_data = TrainingQuestionSerializer(current_question).data

            return Response({"status": True,"message": "Next question fetched successfully.","data": question_data})
        else:
            return Response({"status": True,"message": "All questions completed.","data": []})

class GetNextQuestion(viewsets.ModelViewSet):
    queryset = QuizSession.objects.all()
    serializer_class = QuizSessionSerializer
    lookup_field = 'session_id'
    
    def update(self, request, *args, **kwargs):
        session_id = self.kwargs.get('session_id')
        quiz_session = QuizSession.objects.get(id=session_id)
        question_id = request.data.get('question_id')
        user_answer = request.data.get('user_answer')
        user = request.user
        current_question = TrainingQuestions.objects.get(id=question_id)
        correct_answer = current_question.correct_answer
        is_correct = (user_answer == correct_answer)

        if is_correct:
            quiz_session.score += current_question.marks

        quiz_session.current_question_index += 1
        quiz_session.save()

        questions = quiz_session.quiz.questions.all()
        if quiz_session.current_question_index >= len(questions):
        
            pass_criteria = quiz_session.quiz.pass_criteria
            if quiz_session.score >= pass_criteria:
                quiz_session.status = 'passed'  
                user.is_qualification = True
                user.save()
                quiz_session.quiz.status = True
                quiz_session.quiz.save()

            elif quiz_session.attempts >= 3:
                quiz_session.status = 'failed'
                quiz_session.quiz.status = False 
                quiz_session.quiz.save()

            elif quiz_session.score < pass_criteria and quiz_session.attempts < 3:
                AttemptLog.objects.create(user=user, quiz_session=quiz_session, question_id=question_id)
                quiz_session.status = 'try_again'
                quiz_session.quiz.status = True
                quiz_session.quiz.save()

            quiz_session.completed_at = timezone.now()
            quiz_session.save()

        return Response({
            "status": True,
            "message": "Answer submitted successfully.",
            "is_correct": is_correct,
            "score": quiz_session.score,
            "next_question_index": quiz_session.current_question_index
        })


class UserIdWiseNoOfAttemptsViewSet(viewsets.ModelViewSet):
    queryset = QuizSession.objects.all()
    serializer_class = QuizSessionSerializer

    def create(self, request, *args, **kwargs):
        try:
            user_id = request.data.get('user_id')
            user = CustomUser.objects.get(id=user_id)
            if not user:
                return Response({'status': False,'message': 'User not found'})
            quiz_sessions = AttemptedQuiz.objects.filter(user=user, quiz__status=True)
            serializer = AttemptedQuizSerializer(quiz_sessions, many=True)
            if not quiz_sessions.exists():
                return Response({'status': False, 'message': 'No quiz sessions found for this user'})
            return Response({'status': True, 'message': 'User attempts fetched successfully', 'data': serializer.data})
        except CustomUser.DoesNotExist:return Response({'status': False,'message': 'User not found'})
        except Exception as e:
            return Response({'status': False,'message': 'Something went wrong', 'error': str(e)})
        

class ClassroomUserIdWiseNoOfAttemptsViewSet(viewsets.ModelViewSet):

    def create(self, request, *args, **kwargs):
        try:
            user_id = request.data.get('user_id')
            user = CustomUser.objects.get(id=user_id)
            if not user:
                return Response({'status': False,'message': 'User not found'})
            quiz_sessions = ClassroomAttemptedQuiz.objects.filter(Q(user=user) & (Q(quiz__status=True) | Q(training_quiz__status=True)))
            serializer = ClassroomAttemptedQuizSerializer(quiz_sessions, many=True)
            if not quiz_sessions.exists():
                return Response({'status': False, 'message': 'No quiz sessions found for this user'})
            return Response({'status': True, 'message': 'fetched successfully', 'data': serializer.data})
        except CustomUser.DoesNotExist:return Response({'status': False,'message': 'User not found'})
        except Exception as e:
            return Response({'status': False,'message': 'Something went wrong', 'error': str(e)})

from user_profile.serializers import *
class ClassRoomWiseSelectedUserViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'classroom_id'

    def list(self, request, *args, **kwargs):
        try:
            classroom_id = kwargs.get('classroom_id')

            # Fetch classroom instance
            classroom = ClassroomTraining.objects.get(id=classroom_id)

            # Get users associated with the classroom
            users = classroom.user.all()  # Assuming a ManyToMany relationship with users

            # Serialize users
            serialized_users = CustomUserSerializer(users, many=True).data

            return Response({'status': True, 'message': 'Selected users fetched successfully', 'data': serialized_users})
        except ClassroomTraining.DoesNotExist:
            return Response({'status': False, 'message': 'Classroom not found'})
        except Exception as e:
            return Response({'status': False, 'message': 'Something went wrong', 'error': str(e)})

class FailedUserViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    queryset = QuizSession.objects.all()
    serializer_class = QuizSessionSerializer

    # def list(self, request, *args, **kwargs):
    #     try:
    #         document_id = kwargs.get('document_id')
    #         document = Document.objects.get(id=document_id)
    
    #         # Fetch failed users for the given document
    #         failed_users = QuizSession.objects.filter(
    #             status='failed',
    #             quiz__status=True,
    #             quiz__document=document
    #         ).values_list('user_id', flat=True)  # Extract user IDs only
    
    #         # Fetch assigned users who have taken classroom training
    #         assigned_users = ClassroomTraining.objects.filter(
    #             document=document
    #         ).values_list('user', flat=True)  # Extract assigned user IDs
    
    #         # Filter remaining failed users who haven't completed assessment
    #         remaining_failed_users = CustomUser.objects.filter(
    #             id__in=failed_users
    #         )
    
    #         serializer = CustomUserSerializer(remaining_failed_users, many=True)
    #         return Response({'status': True, 'message': 'Failed users fetched successfully', 'data': serializer.data})
    
    #     except Document.DoesNotExist:
    #         return Response({
    #             'status': False,
    #             'message': 'Document not found'
    #         })
    #     except Exception as e:
    #         return Response({'status': False, 'message': 'Something went wrong', 'error': str(e)})

    def list(self, request, *args, **kwargs):
        try:
            document_id = kwargs.get('document_id')
            document = Document.objects.get(id=document_id)
            failed_users = QuizSession.objects.filter(status='Failed', quiz__status=True,quiz__document = document)
            assigned_users = ClassroomTraining.objects.filter(document=document)
            remaining_failed_users = QuizSession.objects.filter(quiz__document = document,user__in=failed_users.values_list('user', flat=True)).exclude(user__in=assigned_users.values_list('user', flat=True))
            serializer = QuizSessionSerializer(remaining_failed_users, many=True)
            return Response({'status': True, 'message': 'Failed users fetched successfully', 'data': serializer.data})
        except Document.DoesNotExist:
            return Response({
                'status': False,
                'message': 'Document not found'
            })
        except Exception as e:
            return Response({'status': False,'message': 'Something went wrong', 'error': str(e)})

class HrAcknowledgementViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializers_class = HRacnowledgementSerializer

    def create(self, request, *args, **kwargs):
        try:
            if not request.user.groups.filter(name="HR").exists():
                return Response({'status': False, 'message': 'Only HR can create acknowledgements.'})

            user_id = self.kwargs.get('user_id')
            user = CustomUser.objects.get(id=user_id)
            remarks = request.data.get('remarks')
            if not remarks:
                return Response({'status': False, 'message': 'Remarks cannot be empty.'})
            hr_acknowledgement = HRacknowledgement.objects.create(user=user, remarks=remarks)
            serializer = HRacnowledgementSerializer(hr_acknowledgement)
            user.is_induction_complete = True
            user.save()
            return Response({'status': True, 'message': 'Acknowledgement created successfully', 'data': serializer.data})
        except CustomUser.DoesNotExist:
            return Response({'status': False, 'message': 'User not found.'})
        except Exception as e:
            return Response({'status': False, 'message': 'Something went wrong', 'error': str(e)})
        
    def list(self, request, *args, **kwargs):
        try:
            if not request.user.groups.filter(name="HR").exists():
                return Response({'status': False, 'message': 'Only HR can get acknowledgements.'}, status=403)
            
            user_id = self.kwargs.get('user_id')
            user = CustomUser.objects.get(id=user_id)
            hr_acknowledgements = HRacknowledgement.objects.filter(user=user)
            serializer = HRacnowledgementSerializer(hr_acknowledgements, many=True)
            return Response({'status': True, 'message': 'Acknowledgements fetched successfully', 'data': serializer.data})
        except CustomUser.DoesNotExist:
            return Response({'status': False, 'message': 'User not found.'})
        except Exception as e:
            return Response({'status': False, 'message': 'Something went wrong', 'error': str(e)})

import shutil
import time
from datetime import date
class InductionCertificateViewSet(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated] 

    def create(self, request, *args, **kwargs):
        try:
            user_id = self.kwargs.get('user_id')
            user = CustomUser.objects.get(id=user_id)
            username = user.username

            # hr_acknowledgement = HRacknowledgement.objects.filter(user=user).order_by('-id').first()
            # hr_name = hr_acknowledgement.user.get_full_name() if hr_acknowledgement else "HR Manager"

            today_date = date.today().strftime("%d-%m-%Y")  # Format: DD-MM-YYYY

            # 🔹 Define filename based on user & date (to check if already exists)
            filename = f"induction_certificate_{user_id}_{today_date}.pdf"
            file_path = os.path.join(settings.MEDIA_ROOT, 'induction_certificate', filename)
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            # ✅ If the PDF already exists, return the existing file
            if os.path.exists(file_path):
                pdf_file_url = f"{settings.MEDIA_URL}induction_certificate/{filename}"
                full_pdf_file_url = f"{request.scheme}://{request.get_host()}{pdf_file_url}"
                return Response({"status": True, "message": "Certificate already generated.", "data": full_pdf_file_url})

            # 🔹 Prepare user data for the template
            user_data = {
                'username': username,
                'date': today_date
            }
            context = {'users_data': user_data}
            template = get_template('index.html') 
            html_content = template.render(context)

            # 🔹 Find wkhtmltopdf path
            wkhtmltopdf_path = shutil.which("wkhtmltopdf")
            if not wkhtmltopdf_path:
                wkhtmltopdf_path = r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe"  # Windows
                # wkhtmltopdf_path = "/usr/bin/wkhtmltopdf"  # Linux

            config = pdfkit.configuration(wkhtmltopdf=wkhtmltopdf_path)

            # 🔹 Generate PDF
            pdfkit.from_string(html_content, file_path, options={
                'page-size': 'Letter',
                'encoding': 'UTF-8',
                'quiet': '',
                '--enable-local-file-access': ''
            }, configuration=config)

            # ✅ Return the PDF URL
            pdf_file_url = f"{settings.MEDIA_URL}induction_certificate/{filename}"
            full_pdf_file_url = f"{request.scheme}://{request.get_host()}{pdf_file_url}"

            user.is_induction_certificate = True
            user.save()

            return Response({"status": True, "message": "PDF generated successfully", "data": full_pdf_file_url})

        except CustomUser.DoesNotExist:
            return Response({"status": False, "message": "User not found", "data": ""})
        except Exception as e:
            return Response({"status": False, "message": str(e), "data": ""})
        
class TrainerViewSet(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def create(self, request, *args, **kwargs):
        try:
            trainer_name = request.data.get('trainer_name')
            description = request.data.get('description')
            employee_code = request.data.get('employee_code')
            designation = request.data.get('designation')
            department = request.data.get('department')

            if not trainer_name or not description or not employee_code or not department or not designation:
                return Response({"status": False, "message": "all fields are required."})
            
            employee_code_normalized = employee_code.lower()
            if Trainer.objects.filter(employee_code=employee_code_normalized).exists():
                return Response({"status": False, "message": "Trainer with this employee code already exists."})

            trainer = Trainer.objects.create(user=request.user,trainer_name=trainer_name,description=description,employee_code=employee_code,designation=designation,department=department)

            return Response({"status": True, "message": "Trainer created successfully.", "data": TrainerSerializer(trainer).data})

        except Exception as e:
            return Response({"status": False, "message": str(e)})

    def list(self, request, *args, **kwargs):
        try:
            trainers = Trainer.objects.filter(user=request.user)
            serialized_data = TrainerSerializer(trainers, many=True).data

            return Response({"status": True, "message": "Trainer list fetched successfully.", "data": serialized_data})

        except Exception as e:
            return Response({"status": False, "message": str(e)})

    def update(self, request, *args, **kwargs):
        try:
            trainer_id = kwargs.get('trainer_id')
            trainer = Trainer.objects.filter(user=request.user, id=trainer_id).first()

            if not trainer:
                return Response({"status": False, "message": "Trainer not found."},)

            trainer_name = request.data.get('trainer_name')
            description = request.data.get('description')
            employee_code = request.data.get('employee_code')
            designation = request.data.get('designation')
            department = request.data.get('department')

            if trainer_name:
                trainer.trainer_name = trainer_name
            if description:
                trainer.description = description
            if employee_code:
                employee_code_normalized = employee_code.lower()
                if Trainer.objects.filter(employee_code=employee_code_normalized).exists():
                    return Response({"status": False, "message": "Trainer with this employee code already exists."})
                trainer.employee_code = employee_code
            if designation:
                trainer.designation = designation
            if department:
                trainer.department = department
            trainer.save()

            return Response({"status": True, "message": "Trainer updated successfully.", "data": TrainerSerializer(trainer).data})

        except Exception as e:
            return Response({"status": False, "message": str(e)})

    def destroy(self, request, *args, **kwargs):
        try:
            trainer_id = kwargs.get('trainer_id')
            trainer = Trainer.objects.filter(user=request.user, id=trainer_id).first()

            if not trainer:
                return Response({"status": False, "message": "Trainer not found."})

            trainer.delete()

            return Response({"status": True, "message": "Trainer deleted successfully."})

        except Exception as e:
            return Response({"status": False, "message": str(e)})
            
        
class TrainerActiveDeactiveViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    
    def update(self, request, *args, **kwargs):
        try:
            trainer_id = kwargs.get('trainer_id')
            trainer = Trainer.objects.filter(id=trainer_id).first()
            if not trainer:
                return Response({"status": False, "message": "Trainer not found."})
            
            if trainer.is_active:
                trainer.is_active = False
                trainer.save()
                return Response({"status": True, "message": "Trainer deactivated successfully."})
            else:
                trainer.is_active = True
                trainer.save()
                return Response({"status": True, "message": "Trainer activated successfully."})
        except Exception as e:
            return Response({"status": False, "message": str(e)})
    
class ClassroomQuestionViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ClassroomQuestionSerializer
    queryset = ClassroomQuestion.objects.all().order_by('-question_created_at')
    
    def create(self, request, *args, **kwargs):
        try:
            user = self.request.user
            # training = request.data.get('training_id')
            # if not training:
                # return Response({"status": False, "message": "Training ID is required", "data": []})
            classroom_id = request.data.get('classroom_id')
            if not classroom_id:
                return Response({"status": False, "message": "Classroom ID is required", "data": []})

            # training = TrainingCreate.objects.get(id=training)
            # if not training:
                # return Response({"status": False, "message": "Training ID not found", "data": []})
            classroom = ClassroomTraining.objects.get(id=classroom_id)
            if not classroom:
                return Response({"status": False, "message": "Classroom ID not found", "data": []})

            question_type = request.data.get('question_type')
            question_text = request.data.get('question_text')
            options = request.data.get('options', [])
            correct_answer = request.data.get('correct_answer')
            marks = request.data.get('marks')

            if not question_type:
                return Response({"status": False, "message": "Question type is required", "data": []})
            if not question_text:
                return Response({"status": False, "message": "Question text is required", "data": []})
            
            if question_type == 'mcq' and not options:
                return Response({"status": False, "message": "MCQ questions must have options", "data": []})

            if question_type == 'mcq' and correct_answer not in options:
                return Response({"status": False, "message": "MCQ correct answer must be one of the options", "data": []})

            if question_type == 'true_false' and correct_answer not in ['True', 'False']:
                return Response({"status": False, "message": "True/False correct answer must be 'True' or 'False'", "data": []})

            if question_type == 'fill_in_the_blank' and not correct_answer:
                return Response({"status": False, "message": "Fill-in-the-blank questions must have a correct answer", "data": []})
            selected_file_type = request.data.get('selected_file_type')
            selected_file = request.FILES.get('selected_file', None)

            training_question = ClassroomQuestion.objects.create(
                # training=training,
                classroom=classroom,
                
                question_type=question_type,
                question_text=question_text,
                options=options,
                correct_answer=correct_answer,
                marks=marks,
                created_by=user,
                selected_file_type = selected_file_type,
                selected_file=selected_file,
                question_created_at=timezone.now()
            )

            # Serialize and return the response
            serializer = ClassroomQuestionSerializer(training_question, context={'request': request})
            data = serializer.data
            return Response({"status": True, "message": "Training question created successfully", "data": data})

        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e), "data": []})
        

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = ClassroomQuestionSerializer(queryset, many=True, context = {'request': request})
            data = serializer.data
            return Response({"status": True,"message": "Training question list fetched successfully","data": data})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        

class ClassroomIdWiseQuestionsViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ClassroomQuestionSerializer
    queryset = ClassroomQuestion.objects.all().order_by('-question_created_at')
    # lookup_field = 'training_id'
    lookup_field = 'classroom_id'


    def list(self, request, *args, **kwargs):
        try:
            # training_id = self.kwargs.get('training_id')
            classroom_id = self.kwargs.get('classroom_id')
            # if not training_id:
            #     return Response({
            #         "status": False,
            #         "message": "Training ID is required.",
            #         "data": []
            #     })
            if not classroom_id:
                return Response({"status": False, "message": "classroom_id ID is required", "data": []})
                    

            # queryset = self.filter_queryset(self.get_queryset().filter(training_id=training_id))
            queryset = self.filter_queryset(self.get_queryset().filter(classroom=classroom_id))
            serializer = ClassroomQuestionSerializer(queryset, many=True, context={'request': request})
            data = serializer.data

            return Response({
                "status": True,
                "message": "classroom_id question list fetched successfully",
                "data": data
            })

        except Exception as e:
            return Response({
                "status": False,
                "message": "Something went wrong",
                "error": str(e)
            })
        
class ClassroomQuizList(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'classroom_id'
    def list(self, request, *args, **kwargs):
        try:
            classroom_id = kwargs.get('classroom_id')
            classroom = ClassroomTraining.objects.get(id=classroom_id)
            # classroom_quizzes = list(ClassroomQuiz.objects.filter(classroom=classroom))
            document_id = classroom.document.id if classroom.document else None
            # training_quizzes = list(TrainingQuiz.objects.filter(document=document_id))
            if document_id:
                queryset = list(TrainingQuiz.objects.filter(document_id=document_id))
            else:
                queryset = list(ClassroomQuiz.objects.filter(classroom=classroom))
            random.shuffle(queryset)
            serializer = TrainingQuizSerializer(queryset, many=True, context={'request': request})
            return Response({"status": True, "message": "Quizzes retrieved successfully", "data": serializer.data})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e), "data": []})


class ClassroomQuizViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ClassroomQuizSerializer
    queryset = ClassroomQuiz.objects.all().order_by('-created_at')

    def create(self, request, *args, **kwargs):
        try:
            # Extracting data from the request
            user = self.request.user
            classroom_id = request.data.get('classroom_id')
            quiz_name = request.data.get('name')
            pass_criteria = request.data.get('pass_criteria')
            quiz_time = request.data.get('quiz_time')
            total_marks = int(request.data.get('total_marks', 0))
            selected_questions = request.data.get('selected_questions', [])  # For manual quizzes only
            
            pass_criteria = int(pass_criteria) if pass_criteria else 0
            if pass_criteria > total_marks:
                return Response({"status": False, "message": "Pass criteria cannot be greater than total marks", "data": []})

            # Validate required fields
            if not all([classroom_id, quiz_name, pass_criteria, quiz_time, total_marks]):
                return Response({"status": False, "message": "Missing required fields", "data": []})

            try:
                classroom = ClassroomTraining.objects.get(id=classroom_id)
            except ClassroomTraining.DoesNotExist:
                return Response({"status": False, "message": "Document not found", "data": []})

            # Create the new quiz
            quiz = ClassroomQuiz.objects.create(
                quiz_name=quiz_name,
                pass_criteria=pass_criteria,
                quiz_time=quiz_time,
                quiz_type='manual',  # Only allow manual quizzes
                created_by=user,
                classroom=classroom,
            )

            total_marks_accumulated = 0  
            total_questions = 0  

            # Handle manual-type quizzes
            if not selected_questions or not isinstance(selected_questions, list):
                return Response({
                    "status": False,
                    "message": "You must provide a list of selected questions for manual quiz creation.",
                    "data": []
                })

            if isinstance(selected_questions, str) and selected_questions.strip() == "":
                selected_questions = []  # Handle the case where it's an empty string

            if not isinstance(selected_questions, list):
                return Response({
                    "status": False,
                    "message": "selected_questions must be a list.",
                    "data": []
                })

            # Validate that each item in selected_questions is an integer (question ID)
            if any(not isinstance(q, int) for q in selected_questions):
                return Response({
                    "status": False,
                    "message": "Each element in selected_questions must be an integer (question ID).",
                    "data": []
                })

            questions = ClassroomQuestion.objects.filter(
                id__in=selected_questions,
                classroom=classroom,
                status=True
            )

            if len(questions) != len(selected_questions):
                return Response({
                    "status": False,
                    "message": "Some of the selected questions are invalid or inactive.",
                    "data": []
                })

            # Create QuizQuestion for each selected question
            for question in questions:
                if total_marks_accumulated + question.marks > total_marks:
                    return Response({
                        "status": False,
                        "message": f"Adding this question would exceed the total marks. Current total: {total_marks_accumulated}, question marks: {question.marks}",
                        "data": []
                    })
                ClassroomquizQuestion.objects.create(quiz=quiz, question=question, marks=question.marks)
                total_marks_accumulated += question.marks
                total_questions += 1

            # Check for total marks mismatch
            if total_marks_accumulated != total_marks:
                return Response({
                    "status": False,
                    "message": f"Total marks mismatch. The accumulated marks are {total_marks_accumulated}, but the input total_marks was {total_marks}. Please adjust.",
                    "data": []
                })

            # Save the quiz with final total marks and total questions
            quiz.total_marks = total_marks_accumulated
            quiz.total_questions = total_questions
            quiz.save()

            # Return the quiz data
            serializer = ClassroomQuizSerializer(quiz, context={'request': request})
            return Response({"status": True, "message": "Quiz created successfully", "data": serializer.data})

        except IntegrityError as e:
            return Response({"status": False, "message": "Database Integrity Error", "error": str(e), "data": []})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e), "data": []})







    def list(self, request, *args, **kwargs):
        try:
            user = self.request.user
            queryset = ClassroomQuiz.objects.filter(created_by=user).order_by('-created_at')
            serializer = ClassroomQuizSerializer(queryset, many=True, context={'request': request})
            return Response({"status": True, "message": "Quizzes retrieved successfully", "data": serializer.data})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e), "data": []})
        

class ClassroomQuizUpdateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    queryset = ClassroomQuiz.objects.all()
    serializer_class = ClassroomQuizSerializer
    lookup_field = 'quiz_id'

    def update(self, request, *args, **kwargs):
        try:
            quiz_id = kwargs.get('quiz_id')
            quiz = ClassroomQuiz.objects.get(id=quiz_id)

            # Extracting data from the request
            quiz_name = request.data.get('name', quiz.quiz_name)
            pass_criteria = request.data.get('pass_criteria', quiz.pass_criteria)
            quiz_time = request.data.get('quiz_time', quiz.quiz_time)
            total_marks = request.data.get('total_marks', quiz.total_marks)
            selected_questions = request.data.get('selected_questions')  # For manual quizzes only
            # Validate required fields
            if not all([quiz_name, pass_criteria, quiz_time, total_marks]):
                return Response({"status": False, "message": "Missing required fields", "data": []})
            
            # Validate pass criteria
            pass_criteria = int(pass_criteria) if pass_criteria else 0
            if pass_criteria > int(total_marks):
                return Response({"status": False, "message": "Pass criteria cannot be greater than total marks", "data": []})
            
            # Handle manual-type quizzes
            if not selected_questions or not isinstance(selected_questions, list):
                return Response({
                    "status": False,
                    "message": "You must provide a list of selected questions for manual quiz creation.",
                    "data": []
                })
            if isinstance(selected_questions, str) and selected_questions.strip() == "":
                selected_questions = []  # Handle the case where it's an empty string
            if not isinstance(selected_questions, list):
                return Response({
                    "status": False,
                    "message": "selected_questions must be a list.",
                    "data": []
                })
            # Validate that each item in selected_questions is an integer (question ID)
            if any(not isinstance(q, int) for q in selected_questions):
                return Response({
                    "status": False,
                    "message": "Each element in selected_questions must be an integer (question ID).",
                    "data": []
                })
            # Remove all existing QuizQuestion objects
            quiz.questions.all().delete()
            # Create QuizQuestion for each selected question
            total_marks_accumulated = 0
            for question in ClassroomQuestion.objects.filter(id__in=selected_questions, classroom=quiz.classroom, status=True):
                if total_marks_accumulated + question.marks > int(total_marks):
                    return Response({
                        "status": False,
                        "message": f"Adding this question would exceed the total marks. Current total: {total_marks_accumulated}, question marks: {question.marks}",
                        "data": []
                    })
                ClassroomquizQuestion.objects.create(quiz=quiz, question=question, marks=question.marks)
                total_marks_accumulated += question.marks
            # Save the quiz with final total marks and total questions

            # Create the new quiz
            quiz.quiz_name = quiz_name
            quiz.pass_criteria = pass_criteria
            quiz.quiz_time = quiz_time
            # Update the total_questions and total_marks fields
            quiz.total_questions = len(selected_questions)
            quiz.total_marks = total_marks_accumulated  # Update total marks to the accumulated marks of selected questions
            quiz.save()
            
            serializer = ClassroomQuestionSerializer(ClassroomQuestion.objects.filter(classroom=quiz.classroom), many=True, context={'request': request})
            return Response({"status": True, "message": "Quiz updated successfully", "data": serializer.data})
        except ClassroomQuiz.DoesNotExist:
            return Response({"status": False, "message": "Quiz not found"})
        except IntegrityError as e:
            return Response({"status": False, "message": "Database Integrity Error", "error": str(e), "data": []})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e), "data": []})


class ClassroomExamViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    queryset = ClassroomQuizSession.objects.all()
    serializer_class = ClassroomQuizSessionSerializer

    def create(self, request, *args, **kwargs):
        quiz_id = request.data.get('quiz_id')
        quiz = ClassroomQuiz.objects.get(id=quiz_id)
        user = self.request.user 

        quiz_session = ClassroomQuizSession.objects.create(user=user, quiz=quiz)

        return Response({"status": True,"message": "Exam started successfully.","quiz_session_id": quiz_session.id})
    

class ClassroomGetNextQuestionViewSet(viewsets.ModelViewSet):
    queryset = ClassroomQuizSession.objects.all()
    serializer_class = ClassroomQuizSessionSerializer

    def list(self, request, *args, **kwargs):
        # Get the QuizSession instance
        session_id = self.kwargs.get('session_id')
        quiz_session = ClassroomQuizSession.objects.get(id=session_id)
        quiz = quiz_session.quiz

        # Get the questions for this quiz and fetch the next question
        questions = quiz.questions.all()  # Using related_name `questions` on the Quiz model

        # Check if there are more questions to show
        if quiz_session.current_question_index < len(questions):
            current_question = questions[quiz_session.current_question_index]

            # Prepare the data for the next question
            question_data = ClassroomQuestionSerializer(current_question).data

            return Response({"status": True,"message": "Next question fetched successfully.","data": question_data})
        else:
            return Response({"status": True,"message": "All questions completed.","data": []})

class ClassroomUpdateGetNextQuestionViewSet(viewsets.ModelViewSet):
    queryset = ClassroomQuizSession.objects.all()
    serializer_class = ClassroomQuizSessionSerializer
    lookup_field = 'session_id'
    
    def update(self, request, *args, **kwargs):
        session_id = self.kwargs.get('session_id')
        quiz_session = ClassroomQuizSession.objects.get(id=session_id)
        question_id = request.data.get('question_id')
        user_answer = request.data.get('user_answer')
        user = request.user
        current_question = ClassroomQuestion.objects.get(id=question_id)
        correct_answer = current_question.correct_answer
        is_correct = (user_answer == correct_answer)

        if is_correct:
            quiz_session.score += current_question.marks

        quiz_session.current_question_index += 1
        quiz_session.save()

        questions = quiz_session.quiz.questions.all()
        if quiz_session.current_question_index >= len(questions):
        
            pass_criteria = quiz_session.quiz.pass_criteria
            if quiz_session.score >= pass_criteria:
                quiz_session.status = 'passed'  
                
                quiz_session.quiz.status = True
                quiz_session.quiz.save()

            elif quiz_session.attempts >= 3:
                quiz_session.status = 'failed'
                quiz_session.quiz.status = False 
                quiz_session.quiz.save()

            elif quiz_session.score < pass_criteria and quiz_session.attempts < 3:
                AttemptLog.objects.create(user=user, quiz_session=quiz_session, question_id=question_id)
                quiz_session.status = 'try_again'
                quiz_session.quiz.status = True
                quiz_session.quiz.save()

            quiz_session.completed_at = timezone.now()
            quiz_session.save()

        return Response({
            "status": True,
            "message": "Answer submitted successfully.",
            "is_correct": is_correct,
            "score": quiz_session.score,
            "next_question_index": quiz_session.current_question_index
        })


class JobDescriptionCreateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    
    def create(self, request, *args, **kwargs):
        try:
            if not request.user.groups.filter(name='DTC').exists():
                return Response({"status": False, "message": "You don't have permission to create job description."})
            user_id = request.data.get('user_id')
            # job_role_id = request.data.get('job_role_id')
            employee_job_description = request.data.get('employee_job_description')

            # job_role = JobRole.objects.get(id=job_role_id)

            job_description = JobDescription.objects.create(
                user_id=user_id,
                # job_role=job_role,
                employee_job_description=employee_job_description,
                status='pending'
            )
            job_description.user.is_description = True
            job_description.user.is_jr_assign = False
            job_description.user.is_jr_approve = False
            job_description.user.is_tni_generate = False
            job_description.user.is_tni_consent = False
            job_description.user.is_qualification = False
            job_description.user.quiz_attemted = False
            job_description.user.save()
            
            return Response({"status": True, "message": "Job description created successfully"})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        

class SaveJobDescriptionViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    
    def update(self, request, *args, **kwargs):
        try:
            if not request.user.groups.filter(name='DTC').exists():
                return Response({"status": False, "message": "You don't have permission to create job description."})
            user_id = request.data.get('user_id')
            # job_role_id = request.data.get('job_role_id')
            employee_job_description = request.data.get('employee_job_description')

            # job_role = JobRole.objects.get(id=job_role_id)

            job_description = JobDescription.objects.filter(user_id=user_id).order_by('-created_at').first()

            if job_description:
                # Update the existing draft job description
                job_description.employee_job_description = employee_job_description
                job_description.status = 'draft'
                job_description.save()
            else:
                # Create a new draft job description
                job_description = JobDescription.objects.create(
                    user_id=user_id,
                    employee_job_description=employee_job_description,
                    status='draft'
                )
            # job_description.user.is_description = True
            # job_description.user.save()
            
            return Response({"status": True, "message": "Job description saved successfully"})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        

class JobDescriptionList(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    queryset = JobDescription.objects.all()
    serializer_class = JobDescriptionSerializer
    lookup_field = 'user_id'
    
    def list(self, request, *args, **kwargs):
        try:
            user_id = self.kwargs.get('user_id')
            job_description_objects = JobDescription.objects.filter(user=user_id).order_by('-created_at')
            serializer = JobDescriptionSerializer(job_description_objects, many=True)
            return Response({"status": True, "message": "Job descriptions retrieved successfully", "data": serializer.data})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})


class JobDescriptionUpdateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    def update(self, request, *args, **kwargs):
        try:
            if not request.user.groups.filter(name='DTC').exists():
                return Response({"status": False, "message": "You don't have permission to review job description."})

            job_description_id = kwargs.get('job_description_id')

            employee_job_obj = HODRemark.objects.get(id=job_description_id)
            # if employee_job_obj.status != 'send_back':
            #     return Response({"status": False, "message": "You can only update a job description with 'send_back' status"})
            
            job_role_id = request.data.get('job_role_id')
            employee_job_description = request.data.get('employee_job_description')
            job_role_obj = JobRole.objects.get(id=job_role_id)
            employee_job_obj.job_role_id = job_role_id
            employee_job_obj.employee_job_description = employee_job_description
            employee_job_obj.save()
            return Response({"status": True, "message": "Job description updated successfully"})
        
        except JobDescription.DoesNotExist:
            return Response({"status": False, "message": "Job description not found"})
        except CustomUser.DoesNotExist:
            return Response({"status": False, "message": "User not found"})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        
class HODApprovalViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    
    def update(self, request, *args, **kwargs):
        try:
            if not request.user.groups.filter(name='HOD').exists():
                return Response({"status": False, "message": "You don't have permission to review job description."})
            user_id = request.data.get('user_id')
            user = CustomUser.objects.get(id=user_id)
            job_description_id = kwargs.get('job_description_id')
            employee_job_description = JobDescription.objects.get(id=job_description_id)
            description = request.data.get('description')
            remarks = request.data.get('remark')
            status = request.data.get('status')
            if status not in ['approved', 'send_back']:
                return Response({"status": False, "message": "Invalid status"})

            hod_remark = HODRemark.objects.create(
                employee_job_description=employee_job_description,
                remarks=remarks,
                status=status,
                user=user
            )
            if description:
                employee_job_description.employee_job_description = description
                employee_job_description.save()

            employee_job_description.status = status
            employee_job_description.save()
            if status == 'send_back':
                custom_user = employee_job_description.user
                custom_user.is_description = False
                custom_user.save()
            if status == 'approved':
                custom_user = employee_job_description.user
                custom_user.is_jr_approve = True
                custom_user.save()

            return Response({"status": True, "message": "Job description reviewed Successfully"})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        

class AttemptedQuizViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    queryset = AttemptedQuiz.objects.all()
    serializer_class = AttemptedQuizSerializer

    def create(self, request, *args, **kwargs):
        try:
            user_id = request.data.get('user_id')
            document_id = request.data.get('document_id')
            quiz_id = request.data.get('quiz_id')
            questions = request.data.get('questions')
            obtain_marks = request.data.get('obtain_marks')
            total_marks = request.data.get('total_marks')
            total_taken_time = request.data.get('total_taken_time')
            is_pass = request.data.get('is_pass')
            incorrect_questions = request.data.get('incorrect_questions')
            correct_questions = request.data.get('correct_questions')

            user = CustomUser.objects.get(id=user_id)
            document = Document.objects.get(id=document_id)
            quiz = TrainingQuiz.objects.get(id=quiz_id)

            AttemptedQuiz.objects.filter(user=user, document=document).update(training_assesment_attempted=False)
            previous_session = QuizSession.objects.filter(user=user, quiz=quiz).order_by('-id').first()
            assigned_document_version = document.version

            if previous_session:
                previous_major_version = previous_session.document_version.split('.')[0]
                current_major_version = assigned_document_version.split('.')[0]

                # Major version check
                if previous_major_version != current_major_version:
                    user.is_tni_consent = True
                    user.save()
                    user.quiz_attemted = True
                    return Response({"status": True, "message": "Quiz started successfully, new session created due to document version change"})
            
            attempts_count_instance = QuizSession.objects.get_or_create(user=user, quiz=quiz, defaults={"attempts": 0})
            attempts_count = QuizSession.objects.filter(user=user, quiz=quiz).first()

            if attempts_count and attempts_count.attempts >= 3:
                session = Session.objects.filter(user_ids=user).first()
                if not session:
                    return Response({"status": False,"message": "Please attend the classroom session before attempting the quiz."})
                # Ensure the session is completed before allowing new attempt
                session_complete = SessionComplete.objects.filter(session=session, user=user, is_completed=True).first()
                if not session_complete:
                    return Response({
                        "status": False,
                        "message": "You must complete the session before starting the exam again."
                    })
                
            user.quiz_attemted = True
            attempted_quiz = AttemptedQuiz.objects.create(
                user=user,
                document=document,
                quiz=quiz,
                obtain_marks=obtain_marks,
                total_marks=total_marks,
                total_taken_time=total_taken_time,
                is_pass=is_pass,
                training_assesment_attempted = False
            )
            if is_pass:
                quiz_session = QuizSession.objects.get(user=user, quiz=quiz)
                quiz_session.status = 'passed'
                quiz_session.save()
                user.is_qualification = True
                user.save()
            else:
                UserCompleteViewDocument.objects.filter(user=user, document=document).delete()
                user.is_tni_generate = False 
                user.save()
            if obtain_marks:
                quiz_session = QuizSession.objects.get(user=user, quiz=quiz)
                quiz_session.score = obtain_marks
                quiz_session.save()
            attempts_count = AttemptedQuiz.objects.filter(user=user, quiz=quiz).count()
            quiz_session, created = QuizSession.objects.get_or_create(
                user=user,
                quiz=quiz,
                defaults={
                    "attempts": 1,  # First attempt should be 1, not 2
                    "document_version": assigned_document_version,
                }
            )
            if not created and not is_pass:
                quiz_session.attempts += 1
                quiz_session.document_version = assigned_document_version
                quiz_session.save()
                if attempts_count >= 3:
                    quiz_session.status = 'Failed'
                    quiz_session.save()
            
            for question in questions:
                question_id = question.get('question_id')
                question_text = question.get('question_text')
                user_answer = question.get('user_answer')
                correct_answer = question.get('correct_answer')
                attempted_question = AttemptedQuizQuestion.objects.create(
                    attempted_quiz=attempted_quiz,
                    question_id=question_id,
                    question_text=question_text,
                    user_answer=user_answer,
                    correct_answer=correct_answer
                )
            for incorrect in incorrect_questions:
                question_id = incorrect.get('question_id')
                question_text = incorrect.get('question_text')
                user_answer = incorrect.get('user_answer')
                correct_answer = incorrect.get('correct_answer')
                attempted_question = AttemptedIncorrectAnswer.objects.create(
                    attempted_quiz=attempted_quiz,
                    question_id=question_id,
                    question_text=question_text,
                    user_answer=user_answer,
                    correct_answer=correct_answer
                )
            for correct in correct_questions:
                question_id = correct.get('question_id')
                question_text = correct.get('question_text')
                user_answer = correct.get('user_answer')
                correct_answer = correct.get('correct_answer')
                attempted_question = AttemptedCorrectAnswer.objects.create(
                    attempted_quiz=attempted_quiz,
                    question_id=question_id,
                    question_text=question_text,
                    user_answer=user_answer,
                    correct_answer=correct_answer
                )
            user.is_tni_consent = True
            user.quiz_attemted = True
            user.save()
            return Response({"status": True, "message": "Attempted quiz created successfully"})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        

class AttemptedQuizListViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    queryset = AttemptedQuiz.objects.all()
    serializer_class = AttemptedQuizSerializer
    lookup_field = 'user_id'
    
    def list(self, request, *args, **kwargs):
        try:
            user_id = self.kwargs.get('user_id')
            attempted_quiz_objects = AttemptedQuiz.objects.filter(user=user_id)
            serializer = AttemptedQuizSerializer(attempted_quiz_objects, many=True)
            return Response({"status": True, "message": "Attempted quizzes retrieved successfully", "data": serializer.data})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        
    
class UserIdWiseResultViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    queryset = AttemptedQuiz.objects.all()
    serializer_class = AttemptedQuizSerializer

    def list(self, request, *args, **kwargs):
        try:
            user_id = self.kwargs.get('user_id')
            document_id = self.kwargs.get('document_id')

            if not user_id:
                return Response({"status": False, "message": "user_id is required."})
            if not document_id:
                return Response({"status": False, "message": "document_id is required."})

            attempted_quiz_objects = AttemptedQuizQuestion.objects.filter(attempted_quiz__user=user_id, attempted_quiz__document=document_id)
            incorrect_question_obj = AttemptedIncorrectAnswer.objects.filter(attempted_quiz__user=user_id, attempted_quiz__document=document_id)
            correct_question_obj = AttemptedCorrectAnswer.objects.filter(attempted_quiz__user=user_id, attempted_quiz__document=document_id)

            # Serialize the data
            serializer = AttemptedQuizQuestionSerializer(attempted_quiz_objects, many=True)
            incorrect_serializer = AttemptedIncorrectAnswerSerializer(incorrect_question_obj, many=True)
            correct_serializer = AttemptedCorrectAnswerSerializer(correct_question_obj, many=True)

            return Response({
                "status": True, 
                "message": "Attempted quizzes retrieved successfully",
                "questions": serializer.data,
                "incorrect_questions": incorrect_serializer.data, 
                "correct_questions": correct_serializer.data
            })

        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        
from datetime import datetime
import platform
class TrainingCompletionViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request, *args, **kwargs):
        try:
            user_id = self.kwargs.get('user_id')

            if not user_id:
                return Response({"status": False, "message": "user_id is required."})
            user = CustomUser.objects.get(id=user_id)
            username = user.username
            print("username",username)
            current_datetime = datetime.now().strftime("%Y-%m-%d")
            job_assign = JobAssign.objects.filter(user=user).first()
            job_role = job_assign.job_roles.all().values_list('job_role_name', flat=True)
            if not job_assign:
                return Response({"status": False, "message": "Job Assign not found."})
            user_data = {
                'username': username,
                'job_assign': ', '.join(job_role),
                'date_time' : current_datetime,
                'logo': os.path.join(settings.BASE_DIR, 'static', 'training_certificate', 'logo.jpeg')
            }

            context = {'users_data': user_data}
            template = get_template('training_completion_certificate.html') 
            html_content = template.render(context)
            print(html_content, 'ddd')
            timestamp = int(time.time())
            filename = f"training_completion_certificate{timestamp}.pdf"
            file_path = os.path.join(settings.MEDIA_ROOT, 'training_completion_certificate', filename)
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            if platform.system() == "Windows":
                wkhtmltopdf_path = "C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe"
            else:
                wkhtmltopdf_path = shutil.which("wkhtmltopdf") or "/usr/bin/wkhtmltopdf"
            config = pdfkit.configuration(wkhtmltopdf=wkhtmltopdf_path)

            pdfkit.from_string(html_content, file_path, options={
                'page-size': 'Letter',
                'encoding': 'UTF-8',
                'quiet': '',
                '--enable-local-file-access': ''
            }, configuration=config)

            pdf_file_url = f"{settings.MEDIA_URL}training_completion_certificate/{filename}"
            full_pdf_file_url = f"{request.scheme}://{request.get_host()}{pdf_file_url}"
            return Response({"status": True, "message": "PDF generated successfully", "data": full_pdf_file_url})

        except CustomUser.DoesNotExist:
            return Response({"status": False, "message": "User not found", "data": ""})
        except Exception as e:
            return Response({"status": False, "message": str(e), "data": ""})
        
class TrainingAttendanceViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request, *args, **kwargs):
        try:
            document_id = self.kwargs.get('document_id')
            classroom = ClassroomTraining.objects.filter(document=document_id).first()
            session = Session.objects.filter(classroom=classroom).first()
            if not session:
                return Response({"status": False, "message": "Session not found."})
            user_in_session = session.user_ids.all()
            userlist_data = []
            for user in user_in_session:
                userlist_data.append({
                    'name': user.username,
                    'department': user.department.department_name  # Assuming department is a ForeignKey
                    })

            user_data = {
                'training_title': classroom.document.document_title,
                'training_number': classroom.document.document_number,
                'version': classroom.document.version,
                'trainer' : classroom.trainer.trainer_name,
                'start_time' : session.start_date,
                'users' : userlist_data,
            }

            context = {'users_data': user_data}
            template = get_template('training_attendance_sheet.html') 
            html_content = template.render(context)
            print(html_content, 'ddd')
            timestamp = int(time.time())
            filename = f"training_attendance_sheet{timestamp}.pdf"
            file_path = os.path.join(settings.MEDIA_ROOT, 'training_attendance_sheet', filename)
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            if platform.system() == "Windows":
                wkhtmltopdf_path = "C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe"
            else:
                wkhtmltopdf_path = shutil.which("wkhtmltopdf") or "/usr/bin/wkhtmltopdf"
            config = pdfkit.configuration(wkhtmltopdf=wkhtmltopdf_path)

            pdfkit.from_string(html_content, file_path, options={
                'page-size': 'Letter',
                'encoding': 'UTF-8',
                'quiet': '',
                '--enable-local-file-access': ''
            }, configuration=config)

            pdf_file_url = f"{settings.MEDIA_URL}training_attendance_sheet/{filename}"
            full_pdf_file_url = f"{request.scheme}://{request.get_host()}{pdf_file_url}"
            return Response({"status": True, "message": "PDF generated successfully", "data": full_pdf_file_url})

        except CustomUser.DoesNotExist:
            return Response({"status": False, "message": "User not found", "data": ""})
        except Exception as e:
            return Response({"status": False, "message": str(e), "data": ""})
        
class ClassroomTrainingAttendanceViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request, *args, **kwargs):
        try:
            classroom = self.kwargs.get('classroom_id')
            classroom = ClassroomTraining.objects.filter(id=classroom).first()
            session = Session.objects.filter(classroom=classroom).first()
            if not session:
                return Response({"status": False, "message": "Session not found."})
            user_in_session = session.user_ids.all()
            userlist_data = []
            for user in user_in_session:
                userlist_data.append({
                    'name': user.username,
                    'department': user.department.department_name if user.department else "No Department"  # Assuming department is a ForeignKey
                    })

            user_data = {
                    'training_title': classroom.document.document_title if classroom.document else "No Document",
                    'training_number': classroom.document.document_number if classroom.document else "N/A",
                    'version': classroom.document.version if classroom.document else "N/A",
                    'trainer': classroom.trainer.trainer_name if classroom.trainer else "No Trainer",
                    'start_time': session.start_date,
                    'users': userlist_data,
                }

            context = {'users_data': user_data}
            template = get_template('training_attendance_sheet.html') 
            html_content = template.render(context)
            print(html_content, 'ddd')
            timestamp = int(time.time())
            filename = f"training_attendance_sheet{timestamp}.pdf"
            file_path = os.path.join(settings.MEDIA_ROOT, 'training_attendance_sheet', filename)
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            if platform.system() == "Windows":
                wkhtmltopdf_path = "C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe"
            else:
                wkhtmltopdf_path = shutil.which("wkhtmltopdf") or "/usr/bin/wkhtmltopdf"
            config = pdfkit.configuration(wkhtmltopdf=wkhtmltopdf_path)

            pdfkit.from_string(html_content, file_path, options={
                'page-size': 'Letter',
                'encoding': 'UTF-8',
                'quiet': '',
                '--enable-local-file-access': ''
            }, configuration=config)

            pdf_file_url = f"{settings.MEDIA_URL}training_attendance_sheet/{filename}"
            full_pdf_file_url = f"{request.scheme}://{request.get_host()}{pdf_file_url}"
            return Response({"status": True, "message": "PDF generated successfully", "data": full_pdf_file_url})

        except CustomUser.DoesNotExist:
            return Response({"status": False, "message": "User not found", "data": ""})
        except Exception as e:
            return Response({"status": False, "message": str(e), "data": ""})
        
class UserCompleteViewDocumentView(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    queryset =UserCompleteViewDocument.objects.all()
    serializer_class = UserCompleteViewDocumentSerializer

    def create(self, request, *args, **kwargs):
        try:
            user_id = request.data.get('user_id')
            document_id = request.data.get('document_id')
            if not user_id or not document_id:
                return Response({"status": False, "message": "User ID and document ID are required."})
            user = CustomUser.objects.get(id=user_id)
            if not user:
                return Response({"status": False, "message": "User not found."})
            document = Document.objects.get(id=document_id)
            if not document:
                return Response({"status": False, "message": "Document not found."})
            UserCompleteViewDocument.objects.create(user=user, document=document)
            user.is_tni_generate = True
            user.save()
            return Response({"status": True, "message": "User completed the document successfully."})
        except Exception as e:
            return Response({"status": False, "message": str(e)})


from django.shortcuts import get_object_or_404
class DashboardDocumentViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    queryset = Document.objects.all()
    serializer_class = DocumentSerializer

    def list(self, request, *args, **kwargs):
        try:
            user_id = kwargs.get('user_id')
            if not user_id:
                return Response({"status": False, "message": "User ID is required."})
            user = get_object_or_404(CustomUser, id=user_id)
            if not user:
                return Response({"status": False, "message": "User not found."})
            job_role = JobAssign.objects.filter(user=user)
            job_role_ids = job_role.values_list('job_roles', flat=True)
            documents = Document.objects.filter(job_roles__id__in=job_role_ids).distinct()
            total_assign_document = documents.count()
            if total_assign_document == 0:
                return Response({"status": False, "message": "No documents assigned to this user's job roles."})
            
            failed_users = QuizSession.objects.filter(user=user, status='failed', quiz__status=True,quiz__document__in = documents).count()

            passed_users = QuizSession.objects.filter(user=user, status='passed', quiz__status=True,quiz__document__in = documents).count()
            return Response({"status": True, "message": "Documents retrieved successfully", "total_assign_document": total_assign_document, "failed_document": failed_users, "passed_document": passed_users})
        
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})






class ClassroomAttemptedQuizViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    queryset = ClassroomAttemptedQuiz.objects.all()
    serializer_class = AttemptedQuizSerializer

    def create(self, request, *args, **kwargs):
        try:
            user_id = request.data.get('user_id')
            # document_id = request.data.get('document_id')
            classroom_id = request.data.get('classroom_id')
            quiz_id = request.data.get('quiz_id')
            questions = request.data.get('questions')
            obtain_marks = request.data.get('obtain_marks')
            total_marks = request.data.get('total_marks')
            total_taken_time = request.data.get('total_taken_time')
            is_pass = request.data.get('is_pass')
            incorrect_questions = request.data.get('incorrect_questions')
            correct_questions = request.data.get('correct_questions')

            user = CustomUser.objects.get(id=user_id)
            # document = Document.objects.get(id=document_id)
            classroom = ClassroomTraining.objects.get(id=classroom_id)
            quiz = None
            training_quiz = TrainingQuiz.objects.filter(id=quiz_id).first()
            if training_quiz:
                quiz = training_quiz
            else:
                classroom_quiz = ClassroomQuiz.objects.filter(id=quiz_id, classroom=classroom).first()
                if classroom_quiz:
                    quiz = classroom_quiz
            # previous_session = QuizSession.objects.filter(user=user, quiz=quiz).order_by('-id').first()
            # assigned_document_version = document.version

            # if previous_session:
            #     previous_major_version = previous_session.document_version.split('.')[0]
            #     current_major_version = assigned_document_version.split('.')[0]

                # Major version check
                # if previous_major_version != current_major_version:
                #     user.is_tni_consent = True
                #     user.save()
                #     user.quiz_attemted = True
                #     return Response({"status": True, "message": "Quiz started successfully, new session created due to document version change"})
            
            # attempts_count_instance = QuizSession.objects.get_or_create(user=user, quiz=quiz, defaults={"attempts": 0})
            # attempts_count = QuizSession.objects.filter(user=user, quiz=quiz).first()

            # if attempts_count and attempts_count.attempts >= 3:
            #     session = Session.objects.filter(user_ids=user).first()
            #     if not session:
            #         return Response({"status": False,"message": "Please attend the classroom session before attempting the quiz."})
            #     # Ensure the session is completed before allowing new attempt
            #     session_complete = SessionComplete.objects.filter(session=session, user=user, is_completed=True).first()
            #     if not session_complete:
            #         return Response({
            #             "status": False,
            #             "message": "You must complete the session before starting the exam again."
            #         })
                
            # user.quiz_attemted = True
            attempted_quiz = ClassroomAttemptedQuiz.objects.create(
                user=user,
                # document=document,
                classroom=classroom,
                training_quiz=quiz if isinstance(quiz, TrainingQuiz) else None,
                quiz=quiz if isinstance(quiz, ClassroomQuiz) else None,
                obtain_marks=obtain_marks,
                total_marks=total_marks,
                total_taken_time=total_taken_time,
                is_pass=is_pass,
                classroom_attempted=True
            )
            # if is_pass:
            #     quiz_session = QuizSession.objects.get(user=user, quiz=quiz)
            #     quiz_session.status = 'passed'
            #     quiz_session.save()
            #     user.is_qualification = True
            #     user.save()
            # else:
            #     user.is_tni_generate = False 
            #     user.save()
            # if obtain_marks:
            #     quiz_session = QuizSession.objects.get(user=user, quiz=quiz)
            #     quiz_session.score = obtain_marks
            #     quiz_session.save()
            # attempts_count = AttemptedQuiz.objects.filter(user=user, quiz=quiz).count()
            # quiz_session, created = QuizSession.objects.get_or_create(
            #     user=user,
            #     quiz=quiz,
            #     defaults={
            #         "attempts": 1,  # First attempt should be 1, not 2
            #         "document_version": assigned_document_version,
            #     }
            # )
            # if not created and not is_pass:
            #     quiz_session.attempts += 1
            #     quiz_session.document_version = assigned_document_version
            #     quiz_session.save()
            #     if attempts_count >= 3:
            #         quiz_session.status = 'Failed'
            #         quiz_session.save()
            
            for question in questions:
                question_id = question.get('question_id')
                question_text = question.get('question_text')
                user_answer = question.get('user_answer')
                correct_answer = question.get('correct_answer')
                attempted_question = ClassroomAttemptedQuizQuestion.objects.create(
                    attempted_quiz=attempted_quiz,
                    question_id=question_id,
                    question_text=question_text,
                    user_answer=user_answer,
                    correct_answer=correct_answer
                )
            for incorrect in incorrect_questions:
                question_id = incorrect.get('question_id')
                question_text = incorrect.get('question_text')
                user_answer = incorrect.get('user_answer')
                correct_answer = incorrect.get('correct_answer')
                attempted_question = ClassroomAttemptedIncorrectAnswer.objects.create(
                    attempted_quiz=attempted_quiz,
                    question_id=question_id,
                    question_text=question_text,
                    user_answer=user_answer,
                    correct_answer=correct_answer
                )
            for correct in correct_questions:
                question_id = correct.get('question_id')
                question_text = correct.get('question_text')
                user_answer = correct.get('user_answer')
                correct_answer = correct.get('correct_answer')
                attempted_question = ClassroomAttemptedCorrectAnswer.objects.create(
                    attempted_quiz=attempted_quiz,
                    question_id=question_id,
                    question_text=question_text,
                    user_answer=user_answer,
                    correct_answer=correct_answer
                )
            # user.is_tni_consent = True
            # user.quiz_attemted = True
            # classroom_user, created = ClassroomTrainingUser.objects.get_or_create(
            # user=user,
            # classroom_training=classroom
            # )
            # classroom_user.assessment_completed = True
            # classroom_user.save()
            return Response({"status": True, "message": "Attempted quiz created successfully"})
        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)})
        
    
class OnceClassroomAttemptedViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def update(self, request, *args, **kwargs):
        try:
            user = request.user
            classroom_id = request.data.get('classroom_id')
            quiz_id = request.data.get('quiz_id')

            # Validate classroom_id and quiz_id
            if not classroom_id or not quiz_id:
                return Response({"status": False, "message": "classroom_id and quiz_id are required"}, status=400)

            # Fetch ClassroomTraining and ClassroomQuiz objects
            try:
                classroom = ClassroomTraining.objects.get(id=classroom_id)
            except ClassroomTraining.DoesNotExist:
                return Response({"status": False, "message": "Classroom training not found"}, status=404)
            
            document_id = classroom.document.id if classroom.document else None

            if document_id:
                quiz = TrainingQuiz.objects.filter(document_id=document_id, id=quiz_id).first()
                if not quiz:
                    return Response({"status": False, "message": "Quiz not found"})
                attempted_quiz = ClassroomAttemptedQuiz.objects.filter(user=user, classroom=classroom, training_quiz=quiz).first()
            else:
                quiz = ClassroomQuiz.objects.filter(classroom=classroom, id=quiz_id).first()
                if not quiz:
                    return Response({"status": False, "message": "Quiz not found"})
                attempted_quiz = ClassroomAttemptedQuiz.objects.filter(user=user, classroom=classroom, quiz=quiz).first()
            # 🔹 Check if an attempted quiz entry exists
            

            if attempted_quiz:
                attempted_quiz.classroom_attempted = True
                attempted_quiz.save()
            else:
                # 🔹 Create a new record if it doesn't exist
                pass

            return Response({"status": True, "message": "Attempted quiz updated successfully"})

        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)}, status=500)


class OnceTrainingAttemptedViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def update(self, request, *args, **kwargs):
        try:
            user = request.user
            document_id = request.data.get('document_id')
            quiz_id = request.data.get('quiz_id')

            if not document_id or not quiz_id:
                return Response({"status": False, "message": "classroom_id and quiz_id are required"}, status=400)

            # Fetch ClassroomTraining and ClassroomQuiz objects
            try:
                document = Document.objects.get(id=document_id)
            except Document.DoesNotExist:
                return Response({"status": False, "message": "Classroom training not found"}, status=404)

            try:
                quiz = TrainingQuiz.objects.get(id=quiz_id)
            except TrainingQuiz.DoesNotExist:
                return Response({"status": False, "message": "Quiz not found"}, status=404)

            # 🔹 Check if an attempted quiz entry exists
            attempted_quiz = AttemptedQuiz.objects.filter(user=user, document=document, quiz=quiz).first()

            if attempted_quiz:
                # 🔹 If multiple records exist, update all of them
                AttemptedQuiz.objects.filter(user=user, document=document, quiz=quiz).update(training_assesment_attempted=True)
            else:
                # 🔹 Create a new record if it doesn't exist
                pass

            return Response({"status": True, "message": "Attempted quiz updated successfully"})

        except Exception as e:
            return Response({"status": False, "message": "Something went wrong", "error": str(e)}, status=500)


class TrainingWiseUsersViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request, *args, **kwargs):
        try:
            training_id = request.query_params.get('document_id')
            
            if not training_id:
                return Response({"status": False, "message": "Training ID is required"}, status=400)

            training = Document.objects.get(id=training_id)

            assigned_job_roles = training.job_roles.all()

            assigned_users = CustomUser.objects.filter(job_assigns__job_roles__in=assigned_job_roles).distinct()

            users_data = []
            for user in assigned_users:
                assessment_attempted = AttemptedQuiz.objects.filter(document=training, user=user).first()
                users_data.append({
                    "id": user.id,
                    "username": user.username,
                    "training_assesment_attempted": assessment_attempted.training_assesment_attempted if assessment_attempted else False
                })

            return Response({
                "status": True,
                "message": "Users assigned to the training document",
                "data": users_data
            })

        except TrainingCreate.DoesNotExist:
            return Response({"status": False, "message": "Training document not found"}, status=404)

        except Exception as e:
            return Response({"status": False, "message": str(e)}, status=500)
        

class OnOffUserForTrainingViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def update(self, request, *args, **kwargs):
        try:
            user_id = request.data.get('user_id')
            training_id = request.data.get('document_id')
            is_active = request.data.get('is_active')
            is_active = str(is_active).lower() in ["true", "1"]
            user = CustomUser.objects.get(id=user_id)
            training = Document.objects.get(id=training_id)
            attemptquiz = AttemptedQuiz.objects.filter(document=training, user=user).first()
            if not user or not training:
                return Response({"status": False, "message": "User or training document not found"})
            if is_active:
                attemptquiz.training_assesment_attempted = False
                attemptquiz.save()
                UserCompleteViewDocument.objects.filter(user=user, document=training).delete()
                return Response({"status": True, "message": "User activated for training document"})
            else:
                attemptquiz.training_assesment_attempted = True
                attemptquiz.save()
                return Response({"status": True, "message": "User deactivated for training document"})
        except Exception as e:
            return Response({"status": False, "message": str(e)})
        

        
import openpyxl
from io import BytesIO
from openpyxl.utils import get_column_letter
class EmployeeRecordLogExcelView(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request, *args, **kwargs):
        try:
            users = CustomUser.objects.all()

            # Create Excel Workbook and Sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Employee Record Excel"

            # Define Headers
            headers = [
                'Employee Name', 'Designation', 'Department',
                'Training Date', 'Training Name', 'Document Number',
                'Current Version', 'Status', 'Trainer Name'
            ]

            # Add Headers to Sheet
            ws.append(headers)

            # Populate Data Rows
            for user in users:
                department_name = user.department.department_name if user.department else "No Department"

                # Fetch Attempted Quiz Data
                attempted_quizzes = AttemptedQuiz.objects.filter(user=user)
                for attempt in attempted_quizzes:
                    ws.append([
                        user.username,
                        user.designation,
                        department_name,
                        attempt.created_at.strftime("%d-%m-%Y"),
                        attempt.document.document_title if attempt.document else "No Title",
                        attempt.document.document_number if attempt.document else "No Document",
                        attempt.document.version if attempt.document else "No Version",
                        "Passed" if attempt.is_pass else "Failed",
                        "-",  # No trainer for AttemptedQuiz
                    ])

                # Fetch Classroom Attempted Quiz Data
                classroom_attempts = ClassroomAttemptedQuiz.objects.filter(user=user)
                for classroom in classroom_attempts:
                    ws.append([
                        user.username,
                        user.designation,
                        department_name,
                        classroom.created_at.strftime("%d-%m-%Y"),
                        classroom.classroom.classroom_name if classroom.classroom else "No Classroom",
                        classroom.classroom.document.document_number if classroom.classroom and classroom.classroom.document else "No Document",
                        classroom.classroom.document.version if classroom.classroom and classroom.classroom.document else "No Version",
                        "Passed" if classroom.is_pass else "Failed",
                        classroom.classroom.trainer.trainer_name if classroom.classroom and classroom.classroom.trainer else "-",
                    ])

            # Adjust Column Widths
            for col_num in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 20

            # Generate Filename
            timestamp = time.strftime("%d_%m_%Y_%H_%M_%S")
            filename = f"employee_excel_log_{timestamp}.xlsx"
            file_path = os.path.join(settings.MEDIA_ROOT, 'employee_excel_log', filename)

            # Create Directory if Not Exists
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            # Save File
            wb.save(file_path)

            # Build File URL
            file_url = request.build_absolute_uri(settings.MEDIA_URL + f'employee_excel_log/{filename}')

            return Response({"status": True, "message": "Excel report generated successfully.", "data": file_url})

        except Exception as e:
            return Response({"status": False, 'message': 'Something went wrong', 'error': str(e)})

class PendingTrainingReportView(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated]
    
    def list(self, request, *args, **kwargs):
        try:
            log_user = request.user
            document_id = kwargs.get('document_id')
            training = Document.objects.get(id=document_id)
            users = CustomUser.objects.filter(job_assigns__job_roles__in=training.job_roles.all()).distinct()

            all_users_data = []
            for user in users:
                if user.department:
                    department_name = user.department.department_name
                else:
                    department_name = "No Department"
                datestatus = AttemptedQuiz.objects.filter(user=user, document=training).first()
                name = training.document_title if training else None
                version = training.version if training else "No Version"
                doc_status = "Passed" if datestatus and datestatus.is_pass else ("Failed" if datestatus else "Pending")
                document_number = training.document_number
                classroom = ClassroomTraining.objects.filter(document=training).first()
                class_status = ClassroomAttemptedQuiz.objects.filter(user=user, classroom=classroom, is_pass=True).exists()
                if class_status:
                    status = "Passed"
                else:
                    status = doc_status
                user_data  = {
                    'employee_name': user.username,
                    'emp_no': user.employee_number,
                    # 'designation': user.designation,
                    'department_name': department_name,
                    # 'training_date': training_date,
                    'training_name': name,
                    'status': status,
                    'document_number': document_number,
                    'current_version': version,
                    # 'trainer_name': trainer_name,
                }
                # print(user_data)
                all_users_data.append(user_data)
            context = {
                'training_type': "SOP",
                'training_no': training.document_number,
                'training_version': training.version,
                'training_title': training.document_title,
                'users_data': all_users_data,
                'log_user': log_user,
                'current_datetime': timezone.now().strftime('%d-%m-%Y %H:%M:%S'),
            }

            template = get_template('pending_training_report.html')
            html = template.render(context)
            print(html)
            timestamp = int(time.time())
            filename = f"pending_training_report{timestamp}.pdf"
            file_path = os.path.join(settings.MEDIA_ROOT, 'pending_training_report', filename)
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            with open(file_path, 'wb') as output_file:
                pisa_status = pisa.CreatePDF(html, dest=output_file)
            if pisa_status.err:
                return Response({"status": False, "message": "Error occurred while generating PDF", "data": ""})
            
            pdf_file_url = f"{settings.MEDIA_URL}pending_training_report/{filename}"
            full_pdf_file_url = f"{request.scheme}://{request.get_host()}{pdf_file_url}"
            return Response({"status": True, "message": "PDF generated successfully", "data": full_pdf_file_url})

        except Document.DoesNotExist:
            return Response({"status": False, "message": "Document not found", "data": ""})
        except Exception as e:
            return Response({"status": False, "message": str(e), "data": ""})
        

class ClassroomIsPreviewViewSet(viewsets.ModelViewSet):

    def update(self, request, *args, **kwargs):
        try:
            classroom_id = kwargs.get('classroom_id')
            user = request.user
            classroom = ClassroomTraining.objects.filter(id=classroom_id).first()
            if not classroom:
                return Response({"status": False, "message": "Classroom not found or you don't have permission to update it"})
            is_preview = request.data.get('is_preview')
            is_preview = str(is_preview).lower() in ["true", "1"]
            preview_obj, created = IsPreviewForClassroom.objects.update_or_create(
                classroom=classroom, user=user,defaults={"is_preview": is_preview}
            )
            return Response({"status": True, "message": "Classroom preview status updated successfully"})
        except Exception as e:
            return Response({"status": False, "message": str(e)})
        
class ClassroomWithoutAssesmentViewSet(viewsets.ModelViewSet):
    queryset = ClassroomTraining.objects.all()
    serializer_class = ClassroomTrainingSerializer
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request, *args, **kwargs):
        try:
            user_id = kwargs.get('user_id')
            if not user_id:
                return Response({"status": False, "message": "User not found"})
            user = CustomUser.objects.get(id=user_id)
            present_sessions = Attendance.objects.filter(user=user, status=Attendance.PRESENT).values_list('session_id', flat=True)
            classrooms = ClassroomTraining.objects.filter(user=user, is_assesment="Without Assessment", sessions__id__in=present_sessions).distinct()
            serializer = ClassroomTrainingSerializer(classrooms, many=True)
            return Response({"status": True, "message": "Classrooms fetched successfully", "data": serializer.data})
        except Exception as e:
            return Response({"status": False, "message": str(e)})

